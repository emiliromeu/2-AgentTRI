"""Piso 13Y (B5): bateria de conservació total -- PERMANENT.

Recalcula EN INDEPENDÈNCIA (llegint només rebudes/validadas i
apartados/ingressos_validadas + decisions.csv -- MAI important
sumar.py/informe.py/app.py, un oracle que crida el codi que vol
auditar no serveix de res) els cubs per tipus d'IVA, Exempta, Base+IVA,
Σ retencions i RESULTAT que sumar.py escriu a l'Excel, per a CADA
trimestre, i els compara CEL·LA A CEL·LA (tolerància 0,01 €) contra el
fitxer ja generat. També comprova, fila a fila del DETALL, que
TOTAL FRA. = Base+IVA − Retenció -- una consistència pura dins del
mateix Excel.

Piso 14C: trampa de no-suma PERMANENT -- per cada factura, Base+IVA
ha de ser EXACTAMENT la Σ de les seves pròpies línies (base+quota),
tant al JSON validat (cuadre_esperado ja ho garanteix indirectament,
però aquí es comprova explícita i directa per fitxa) com llegint-ho
del DETALL de l'Excel mateix (agrupant les files que pertanyen a la
mateixa factura, ja que Base+IVA només s'escriu a la primera línia
del grup fusionat -- Piso 14). És impossible per construcció que
Base+IVA superi la Σ de les seves línies: si no hi ha manera de
quadrar-ho, validar.py deixa la fitxa PENDENT, mai la deixa passar.

Les correccions de correccions.csv NO calen llegir-les a part: quan
una factura arriba a validadas/*.json ja les porta aplicades (validar.py
les aplica abans d'escriure-hi) -- per això aquest script només
necessita validadas/*.json + decisions.csv per tenir la mateixa
informació que sumar.py fa servir per sumar.

Busca les etiquetes conegudes ("21", "10", "EXEMPTA", "ALTRES", "TOTAL",
"Σ RETENCIONS SUPORTADES", "TOTAL RETENCIONS", "RESULTAT IVA...") en
comptes d'assumir files fixes -- el layout de sumar.py pot canviar de
fila d'una execució a l'altra sense que aquest script hagi de canviar.

Regla CLAUDE.md (Piso 13Y): cap piso futur toca sumar.py sense passar
aquest script contra els 4 clients reals abans de commitejar.

Ús:
    python3 verificar_conservacio.py <carpeta_client> [<carpeta2> ...]
    python3 verificar_conservacio.py            # tots els clients
"""
import csv
import json
import os
import sys

import openpyxl

TIPOS_IVA = [0, 4, 5, 10, 12, 21]
TOLERANCIA = 0.01


def leer_clientes():
    ruta = "clientes/clientes.csv"
    if not os.path.exists(ruta):
        return []
    with open(ruta, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def cargar_validadas(carpeta):
    facturas = []
    if not os.path.isdir(carpeta):
        return facturas
    for nombre in sorted(os.listdir(carpeta)):
        if not nombre.lower().endswith(".json"):
            continue
        with open(os.path.join(carpeta, nombre), encoding="utf-8") as f:
            facturas.append((nombre, json.load(f)))
    return facturas


def cargar_decisiones(carpeta_cliente):
    """Còpia INDEPENDENT de la mateixa gramàtica (Piso 13M) que
    app.py/sumar.py/informe.py -- última fila per arxiu, "revertir"
    treu l'arxiu del diccionari."""
    ruta = f"{carpeta_cliente}/decisions.csv"
    decisiones = {}
    if not os.path.exists(ruta):
        return decisiones
    with open(ruta, encoding="utf-8") as f:
        for fila in csv.DictReader(f):
            archivo = fila.get("archivo")
            if not archivo:
                continue
            if fila.get("accion") == "revertir":
                decisiones.pop(archivo, None)
            else:
                decisiones[archivo] = fila
    return decisiones


def estat_efectiu(decision):
    return decision.get("accion") if decision else None


def trimestre_de(fecha):
    if not fecha:
        return None
    try:
        mes = int(fecha[5:7])
    except (ValueError, IndexError):
        return None
    return f"{(mes - 1) // 3 + 1}T"


def cuadre_esperado(facturas, decisiones):
    """Base/quota per tipus (incl. "otros"/"exenta"), total i retenció
    de les factures que COMPTEN -- estado==OK o decisió aprovar, mai
    descartar. Mateixa regla que sumar_bloque, calculada aquí en
    independència total.

    Piso 14B: exenta és un camp DE LÍNIA (validar.py ja el resol i
    deriva el de factura) -- una factura mixta reparteix cada línia
    al seu cub, igual que sumar_bloque."""
    sumas = {t: {"base": 0.0, "cuota": 0.0} for t in TIPOS_IVA}
    sumas["otros"] = {"base": 0.0, "cuota": 0.0}
    sumas["exenta"] = {"base": 0.0, "cuota": 0.0}
    total = 0.0
    retencion = 0.0
    for nombre, datos in facturas:
        decision = decisiones.get(nombre)
        if estat_efectiu(decision) == "descartar":
            continue
        cuenta = datos.get("estado") == "OK" or estat_efectiu(decision) == "aprovar"
        if not cuenta:
            continue
        for linea in datos.get("lineas_iva") or []:
            tipo = linea.get("tipo_iva")
            clave = "exenta" if linea.get("exenta") else (tipo if tipo in TIPOS_IVA else "otros")
            sumas[clave]["base"] += linea.get("base") or 0
            sumas[clave]["cuota"] += linea.get("cuota") or 0
        total += datos.get("total") or 0
        retencion += datos.get("retencion_cuota") or 0
    return sumas, round(total, 2), round(retencion, 2)


def verificar_no_suma_json(facturas):
    """Piso 14C: per cada factura que la MÀQUINA ha validat OK (mai les
    que només compten per un Aprovar manual -- aquest és precisament
    l'escapatoria humana per fer passar una fitxa que l'aritmètica NO
    quadra, per exemple un paper amb un error d'impremta que algú ha
    verificat a mà; "impossible per construcció" es refereix a la
    xarxa de validació, mai a una decisió humana que la sobreescriu a
    consciència, amb qui/data/nota al llibre major), Base+IVA
    (datos["total"], ja normalitzat per validar.py si calia) ha de ser
    EXACTAMENT la Σ de les seves pròpies línies -- independent de
    l'Excel, directe sobre validadas/*.json. Retorna la llista de
    violacions (nombre, esperat, real) -- buida si tot quadra."""
    violacions = []
    for nombre, datos in facturas:
        if datos.get("estado") != "OK":
            continue
        suma_lineas = round(
            sum((l.get("base") or 0) + (l.get("cuota") or 0) for l in (datos.get("lineas_iva") or [])), 2,
        )
        total = round(datos.get("total") or 0, 2)
        if abs(suma_lineas - total) > TOLERANCIA:
            violacions.append((nombre, suma_lineas, total))
    return violacions


SECCIONS_QUE_TANQUEN = {
    "RESULTAT DEL TRIMESTRE", "DESPESES", "INGRESSOS",
    "DETALL DESPESES", "DETALL INGRESSOS",
    "PENDENT DE REVISIÓ", "PENDENTS — NO SUMEN", "DESCARTATS",
}


def tipo_clave_de(etiqueta):
    if isinstance(etiqueta, (int, float)) and not isinstance(etiqueta, bool):
        return int(etiqueta)
    if etiqueta == "ALTRES":
        return "otros"
    if etiqueta == "EXEMPTA":
        return "exenta"
    return None


def leer_hoja_trimestre(ws):
    """Escaneja el full buscant les etiquetes conegudes -- mai assumeix
    files fixes. Retorna els cubs/totals de RESULTAT/DESPESES/INGRESSOS
    i, per cada fila de DETALL amb un Base+IVA, (base_iva, retenció,
    total_fra) per verificar TOTAL FRA. = Base+IVA - Retenció, i
    (num_factura, base_iva, Σ de les seves pròpies línies) per a la
    trampa de no-suma (Piso 14C)."""
    valores = {
        "resultat_g": {}, "resultat_i": {}, "resultat_iva": None,
        "despeses_base": {}, "despeses_cuota": {}, "despeses_total": None,
        "despeses_retencio": None, "despeses_total_fra": None,
        "ingressos_base": {}, "ingressos_cuota": {}, "ingressos_total": None,
        "ingressos_retencio": None, "ingressos_total_fra": None,
        "filas_liquid": [],
        "filas_no_suma": [],
    }
    seccion = None
    mapa_cols_detall = None
    # Piso 14C: la trampa de no-suma NOMÉS te sentit dins "FACTURES QUE
    # SUMEN" -- "PENDENTS -- NO SUMEN" son precisament les que NO
    # quadren (aquest es sovint el seu propi motiu de quedar PENDENT),
    # mai una violació.
    en_que_suman = False
    # Piso 14C: acumulador del grup de files (fusionades, Piso 14) de
    # LA FACTURA EN CURS -- Base+IVA nomes s'escriu a la primera línia
    # del grup, així que cal sumar Base+Quota de totes les seves
    # línies abans de poder comparar-ho.
    grup_actual = None

    def tancar_grup():
        # Piso 14C: un Aprovar manual pot fer comptar una fitxa que
        # l'aritmètica NO quadra a consciència (decisions.csv en
        # guarda qui/data/nota) -- exactament l'escapatoria humana que
        # "impossible per construcció" no pretén tapar, nomes la xarxa
        # automàtica de validar.py.
        if grup_actual is not None and "aprovat manualment" not in (grup_actual["estat"] or ""):
            valores["filas_no_suma"].append((
                grup_actual["num_factura"], grup_actual["base_iva"], round(grup_actual["suma_linies"], 2),
            ))

    for fila in ws.iter_rows():
        etiqueta = fila[0].value

        if etiqueta == "FACTURES QUE SUMEN":
            en_que_suman = True
        elif etiqueta in SECCIONS_QUE_TANQUEN:
            en_que_suman = False

        if etiqueta in SECCIONS_QUE_TANQUEN:
            tancar_grup()
            grup_actual = None
            seccion = {
                "RESULTAT DEL TRIMESTRE": "resultat", "DESPESES": "despeses", "INGRESSOS": "ingressos",
            }.get(etiqueta)
            mapa_cols_detall = None
            continue
        # Piso 14C bugfix: una línia de continuació d'una factura fusionada
        # (Piso 14) porta Data (columna A) en blanc pero SI que porta Base/
        # Quota propis -- saltar-la aquí abans d'arribar al bloc de
        # mapa_cols_detall feia que el grup namés sumés la primera línia
        # (falsos positius de "no-suma" a qualsevol factura multi-línia).
        # Només saltem si NO som dins d'una taula de DETALL.
        if etiqueta is None and mapa_cols_detall is None:
            continue

        # DETALL: la fila de capçalera ("Data", "Núm. factura"...) marca
        # per etiqueta on és cada columna -- es rellegeix a cada bloc
        # (FACTURES QUE SUMEN i PENDENTS -- NO SUMEN en tenen una cada un)
        # perquè cap posició de columna es dona per fixa.
        if etiqueta == "Data":
            tancar_grup()
            grup_actual = None
            mapa_cols_detall = {cel.value: idx for idx, cel in enumerate(fila, start=1) if cel.value}
            continue
        if mapa_cols_detall is not None:
            col_base_iva = mapa_cols_detall.get("Base + IVA (€)")
            col_total_fra = mapa_cols_detall.get("TOTAL FRA. (€)")
            col_retencio = mapa_cols_detall.get("Retenció (€)")
            col_base = mapa_cols_detall.get("Base")
            col_quota = mapa_cols_detall.get("Quota")
            col_estat = mapa_cols_detall.get("Estat")
            if col_base_iva and col_total_fra:
                valor_base_iva = fila[col_base_iva - 1].value
                if valor_base_iva is not None:
                    # Nova factura -- tanca el grup anterior i comença.
                    tancar_grup()
                    valor_retencio = (fila[col_retencio - 1].value if col_retencio else None) or 0
                    valor_total_fra = fila[col_total_fra - 1].value
                    valores["filas_liquid"].append((fila[1].value, valor_base_iva, valor_retencio, valor_total_fra))
                    if en_que_suman:
                        valor_estat = fila[col_estat - 1].value if col_estat else None
                        grup_actual = {
                            "num_factura": fila[1].value, "base_iva": valor_base_iva,
                            "suma_linies": 0.0, "estat": valor_estat,
                        }
                    else:
                        grup_actual = None
            if grup_actual is not None and col_base and col_quota:
                grup_actual["suma_linies"] += (fila[col_base - 1].value or 0) + (fila[col_quota - 1].value or 0)

        tipo_clave = tipo_clave_de(etiqueta)
        if seccion == "resultat" and tipo_clave is not None:
            valores["resultat_g"][tipo_clave] = fila[1].value or 0
            valores["resultat_i"][tipo_clave] = fila[2].value or 0
        elif seccion == "resultat" and etiqueta == "RESULTAT IVA (repercutit - suportat)":
            valores["resultat_iva"] = fila[2].value or 0
        elif seccion == "despeses" and tipo_clave is not None:
            valores["despeses_base"][tipo_clave] = fila[1].value or 0
            valores["despeses_cuota"][tipo_clave] = fila[2].value or 0
        elif seccion == "despeses" and etiqueta == "TOTAL":
            valores["despeses_total"] = fila[3].value or 0
        elif seccion == "despeses" and etiqueta in ("Σ RETENCIONS SUPORTADES", "TOTAL RETENCIONS"):
            valores["despeses_retencio"] = fila[3].value or 0
        elif seccion == "despeses" and etiqueta == "Σ TOTAL FRA.":
            valores["despeses_total_fra"] = fila[3].value or 0
        elif seccion == "ingressos" and tipo_clave is not None:
            valores["ingressos_base"][tipo_clave] = fila[1].value or 0
            valores["ingressos_cuota"][tipo_clave] = fila[2].value or 0
        elif seccion == "ingressos" and etiqueta == "TOTAL":
            valores["ingressos_total"] = fila[3].value or 0
        elif seccion == "ingressos" and etiqueta in ("Σ RETENCIONS SUPORTADES", "TOTAL RETENCIONS"):
            valores["ingressos_retencio"] = fila[3].value or 0
        elif seccion == "ingressos" and etiqueta == "Σ TOTAL FRA.":
            valores["ingressos_total_fra"] = fila[3].value or 0

    tancar_grup()
    return valores


def comparar(etiqueta, esperado, real, detalls=""):
    if abs((real or 0) - (esperado or 0)) <= TOLERANCIA:
        return True
    print(f"  ✗ {etiqueta}: esperat {esperado}, real {real}{detalls}")
    return False


def verificar_cliente(carpeta):
    carpeta_cliente = f"clientes/{carpeta}"
    ruta_xlsx = f"{carpeta_cliente}/sumatorios_2026.xlsx"
    if not os.path.exists(ruta_xlsx):
        print(f"{carpeta}: sense Excel generat, es salta.")
        return True

    gastos = cargar_validadas(f"{carpeta_cliente}/rebudes/validadas")
    ingresos = cargar_validadas(f"{carpeta_cliente}/apartados/ingressos_validadas")
    decisiones = cargar_decisiones(carpeta_cliente)

    trimestres = {}
    for nombre, datos in gastos:
        t = trimestre_de(datos.get("fecha_factura"))
        if t:
            trimestres.setdefault(t, {"gastos": [], "ingresos": []})["gastos"].append((nombre, datos))
    for nombre, datos in ingresos:
        t = trimestre_de(datos.get("fecha_factura"))
        if t:
            trimestres.setdefault(t, {"gastos": [], "ingresos": []})["ingresos"].append((nombre, datos))

    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)
    tot_ok = True

    for trimestre, datos_t in sorted(trimestres.items()):
        if trimestre not in wb.sheetnames:
            print(f"{carpeta} / {trimestre}: ✗ no hi ha full a l'Excel per a aquest trimestre")
            tot_ok = False
            continue

        esperado_g, total_g, retencio_g = cuadre_esperado(datos_t["gastos"], decisiones)
        esperado_i, total_i, retencio_i = cuadre_esperado(datos_t["ingresos"], decisiones)
        resultat_iva_esperat = round(
            sum(esperado_i[t]["cuota"] for t in esperado_i) - sum(esperado_g[t]["cuota"] for t in esperado_g), 2
        )

        real = leer_hoja_trimestre(wb[trimestre])
        ok = True

        for tipo in real["despeses_base"]:
            ok &= comparar(f"{trimestre} DESPESES base {tipo}", esperado_g.get(tipo, {}).get("base"), real["despeses_base"][tipo])
        for tipo in real["despeses_cuota"]:
            ok &= comparar(f"{trimestre} DESPESES quota {tipo}", esperado_g.get(tipo, {}).get("cuota"), real["despeses_cuota"][tipo])
        ok &= comparar(f"{trimestre} DESPESES Base+IVA", total_g, real["despeses_total"])
        ok &= comparar(f"{trimestre} DESPESES retenció", retencio_g, real["despeses_retencio"])
        ok &= comparar(f"{trimestre} DESPESES Σ TOTAL FRA.", round(total_g - retencio_g, 2), real["despeses_total_fra"])

        for tipo in real["ingressos_base"]:
            ok &= comparar(f"{trimestre} INGRESSOS base {tipo}", esperado_i.get(tipo, {}).get("base"), real["ingressos_base"][tipo])
        for tipo in real["ingressos_cuota"]:
            ok &= comparar(f"{trimestre} INGRESSOS quota {tipo}", esperado_i.get(tipo, {}).get("cuota"), real["ingressos_cuota"][tipo])
        ok &= comparar(f"{trimestre} INGRESSOS Base+IVA", total_i, real["ingressos_total"])
        ok &= comparar(f"{trimestre} INGRESSOS retenció", retencio_i, real["ingressos_retencio"])
        ok &= comparar(f"{trimestre} INGRESSOS Σ TOTAL FRA.", round(total_i - retencio_i, 2), real["ingressos_total_fra"])

        ok &= comparar(f"{trimestre} RESULTAT IVA", resultat_iva_esperat, real["resultat_iva"])

        for num_factura, base_iva, retencio, total_fra in real["filas_liquid"]:
            esperat_total_fra = round((base_iva or 0) - (retencio or 0), 2)
            ok &= comparar(f"{trimestre} TOTAL FRA. de factura {num_factura}", esperat_total_fra, total_fra)

        # Piso 14C: trampa de no-suma -- Base+IVA HA DE SER exactament
        # la Σ de les seves pròpies línies, llegit del mateix Excel
        # (mai pot ser-hi més gran: si no quadrés, validar.py l'hauria
        # deixat PENDENT i mai hauria arribat a comptar-se aquí).
        for num_factura, base_iva, suma_linies in real["filas_no_suma"]:
            ok &= comparar(f"{trimestre} no-suma de factura {num_factura}", suma_linies, base_iva)

        if ok:
            print(f"{carpeta} / {trimestre}: ✓ ({len(real['filas_liquid'])} files de detall comprovades)")
        tot_ok &= ok

    # Piso 14C: trampa de no-suma independent del JSON (mai llegeix
    # l'Excel) -- per a CADA factura que compta, Base+IVA (datos.total,
    # ja normalitzat si calia) ha de ser EXACTAMENT la Σ de les seves
    # línies.
    violacions = verificar_no_suma_json(gastos) + verificar_no_suma_json(ingresos)
    for nombre, esperat, real_total in violacions:
        print(f"  ✗ {carpeta} no-suma JSON de {nombre}: Σlínies={esperat}, total={real_total}")
        tot_ok = False

    return tot_ok


def main():
    carpetas = sys.argv[1:]
    if not carpetas:
        carpetas = [f["carpeta"] for f in leer_clientes()]

    tot_ok = True
    for carpeta in carpetas:
        tot_ok &= verificar_cliente(carpeta)

    print()
    print("CONSERVACIÓ TOTAL: ✓ TOT QUADRA" if tot_ok else "CONSERVACIÓ TOTAL: ✗ HI HA DIVERGÈNCIES -- veure detall a dalt")
    sys.exit(0 if tot_ok else 1)


if __name__ == "__main__":
    main()
