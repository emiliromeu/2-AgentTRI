"""Agrupa las facturas validadas por trimestre y tipo de IVA, y escribe
un Excel de sumatorios por cliente.

Piso 5: sin llamadas a la API, corre gratis. Lee rebudes/validadas/
(GASTOS) y apartados/ingressos_validadas/ (INGRESOS) de cada cliente --
son fixtures, nunca se tocan. Cada run reescribe el Excel entero: no
hay nada que idempotizar, es gratis y siempre tiene que reflejar el
ultimo estado de las validadas, no el de la vez anterior.

Piso 5B: viste el Excel -- bloque de titulo por hoja, colores suaves,
formato de moneda, bordes en TOTAL/retenciones, anchos de columna y
ajuste de texto en Motivos. La estructura de datos no cambia.

Piso 6A: hoja final "AVISOS" por cliente (regla 9 de CLAUDE.md --
nada muere en silencio tampoco en el enrutado): albaranes apartados,
paginas de ruido leidas de los manifiestos de trocear.py, y avisos de
consistencia (mismo proveedor, mas de un tipo de IVA). Solo informa,
no cambia ningun estado OK/REVISAR.

Piso 6B: encontrar_original tambien busca en subcarpetas por proveedor
(clientes con facturas de compra organizadas asi, ej. davinstal), y el
origen de las emitidas es personalizable por cliente.

Piso 7: DETALLE GASTOS/INGRESOS por hoja de trimestre -- una fila por
linea de IVA de cada factura (OK y REVISAR), para auditar de donde
sale cada suma. No cambia ningun total existente.

Piso 8: todo el texto visible pasa a catala (los motivos de validar.py
se traducen aqui, en la presentacion -- validar.py no se toca). Nueva
fila RESULTAT IVA (calculo derivado de las Sigma cuota ya existentes).
Relleno verde suave para las filas OK (antes solo REVISAR llevaba
relleno). Ningun total cambia: solo presentacion y calculo derivado.

Piso 9.1: RESULTAT DEL TRIMESTRE pasa a tabla propia al inicio de cada
hoja (sustituye la linea suelta del piso 8). Etiqueta ABONAMENT para
total negativo. Dos verificaciones puras (letra de NIF, cuadre de
retencion) que solo avisan, nunca cambian estado ni suma. Seccion
ERRORS en AVISOS -- archivos presentes sin ficha extraida, motivo
generico porque extraer_todas.py no persiste la razon del fallo.

Piso 9.2: decisions.csv (archivo,accion,nota,qui,data) -- 'aprovar'
hace que una REVISAR cuente como OK; 'descartar' la saca de toda suma
y la manda a un bloque DESCARTATS propio. Sin decisions.csv, o vacio,
el comportamiento es identico al de antes de este piso.

Piso 9.3: encontrar_original tambien busca en "procesadas" (antes
excluida). verificar_enlaces_excel() reabre el Excel ya escrito y
distingue "trencat" (bug de codigo, no deberia pasar nunca) de
"corrupte" (archivo presente pero vacio o truncado -- problema de
datos/sincronizacion, ningun cambio de codigo lo arregla). El motivo
de un ERROR distingue este caso del generico cuando se detecta.

Piso 11B: distintivo "CORREGIT" en DETALLE (mismo patron aditivo que
ABONAMENT/[AVÍS]) -- si la ficha trae camps_corregits (escrito por
validar.py, unica cirugia del piso), se anexa al texto del estado con
el detalle antic->nou. Ninguna suma ni logica cambia aqui.

Piso 12A: los totales (RESULTAT, TOTAL de bloque, TOTAL RETENCIONS) pasan
de valores Python ya redondeados a formulas Excel =SUM() reales, para
que Emili pueda ver -y en su caso retocar- de donde sale cada numero
directamente en el libro. Para que esas formulas puedan apuntar a un
rango contiguo, DETALL se reparte en dos bloques por flujo: "FACTURES
QUE SUMEN" (las mismas que hoy cuentan en sumar_bloque, agrupadas por
tipus d'IVA para que cada tipo ocupe un rango de filas seguido) y
"PENDENTS -- NO SUMEN" (las REVISAR sin decision, en taronja, con el
mismo detalle de linias, para poder auditar sin sumar). Las DESCARTADES
ya no se repiten en DETALL -- tienen su propio bloc DESCARTATS con
nota/qui/data. Nueva columna "Retenció" (una fila por factura, no por
linia, para no duplicar la retencio si hay mas de un tipus d'IVA).
Como RESULTAT/DESPESES/INGRESSOS se escriben ANTES que DETALL en la
hoja pero sus formulas necesitan saber en que filas exactas cae cada
tipo dentro de DETALL, cada hoja se escribe en tres pasadas: (1)
etiquetas con las celdas numericas en blanco, recordando su posicion;
(2) PENDENT DE REVISIO y DESCARTATS igual que siempre; (3) DETALL,
anotando el rango de filas de cada tipo y del bloque entero mientras
se escribe. Con esos rangos ya conocidos se rellenan las celdas en
blanco del paso 1 con texto de formula (openpyxl nunca la calcula,
solo Excel al abrir el archivo). Verificacion nueva: verificar_formulas
reabre el Excel ya guardado, reparsa cada formula con una regex,
confirma que el rango coincide exactamente con el anotado y recalcula
en Python el mismo valor a partir de las celdas planas de DETALL,
comparando al centimo contra lo que sumar_bloque ya calculo en memoria.

Piso 13J: diseccion forense (Piso 13I) encontro la causa real de que
Excel 2007 pida "reparar" un archivo recien generado -- openpyxl
escribe toda celda de formula con <v /> vacio (confirmado en su propio
codigo fuente), y una celda que se declara numerica por defecto con
valor vacio es lo que el 2007 rechaza al cargar. inyectar_valores_cacheados
reabre el .xlsx justo despues de guardarlo e injecta ahi el valor que
el propio codigo ya calculo (el mismo que compara verificar_formulas),
porque openpyxl no ofrece ninguna forma de escribirlo por su API.

Piso 13K: la columna 3 de DETALL (antes siempre "proveedor") pasa a
ser la contrapart que ya calcula validar.py -- en DESPESES sigue
diciendose "Proveïdor" (da lo mismo que antes, el cliente ahi es
siempre receptor); en INGRESSOS pasa a decir "Client" y muestra el
comprador real, no el propio cliente cuando el es quien emite.

Piso 13J: bug encontrado verificando migrar_lot.py -- origen_ingressos
ya lleva carpeta_cliente incluido (donde se define, mas arriba en el
bucle principal), pero se volvia a anteponer otra vez en pendientes/
descartados/DETALL INGRESSOS, rompiendo la ruta para encontrar_original.
No era solo un caso latente (ingres PENDENT/DESCARTAT): afectaba a
CUALQUIER hyperlink de "Núm. factura" en DETALL INGRESSOS de CUALQUIER
cliente -- confirmado al arreglarlo, el numero de enllaços verificats
subio de golpe en todos los clientes con ingressos (antes esas celdas
nunca tenian enllaç, en silenci, des de sempre).

Piso 13O: petició d'una revisora -- les dates de l'Excel (columna
"Data" del DETALL, i la cel·la "Generat el..." de la capçalera,
repetida a la fulla AVISOS) passen de text ISO a objecte datetime
real amb number_format dd/mm/yyyy -- Excel ja les tracta com a DATES
de veritat (ordenables, filtrables, llestes per a Geyce V2). Nomes
aquest fitxer canvia -- app.py/informe.py/JSON/CSV interns es queden
en ISO, correcte com estava.
"""

import csv
import json
import os
import re
import time
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import quote, urlparse
from urllib.request import url2pathname

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Piso 13B: ancla el cwd a la carpeta del propio script -- ver trocear.py.
RAIZ = Path(__file__).resolve().parent
os.chdir(RAIZ)

TIPOS_IVA = [0, 4, 5, 10, 12, 21]
EXTENSIONES_ORIGINAL = (".pdf", ".jpg", ".jpeg", ".png")

# Mismo criterio que en extraer_todas.py: algunos clientes ya tenian su
# carpeta de facturas emitidas organizada antes de este pipeline, asi
# que el enlace del original tiene que apuntar ahi en vez de al
# apartados/ingressos/ por defecto.
RUTAS_ORIGEN_INGRESSOS_PERSONALIZADAS = {"davinstal": "Emeses/davinstal"}

SUBCARPETAS_RESERVADAS = {"extraidas", "validadas", "procesadas", "lotes_escaneados", "lotes_procesados"}
# Piso 9.3: para buscar un ORIGINAL (no para detectar errores), "procesadas"
# no se excluye -- ahi es donde un PDF ya tratado puede haberse movido
# (regla 5, run idempotente), y sigue siendo su ubicacion legitima.
SUBCARPETAS_NO_ORIGINALES = {"extraidas", "validadas", "lotes_escaneados", "lotes_procesados"}

FORMATO_MONEDA = '#,##0.00" €"'
FORMATO_PORCENTAJE = '0"%"'

FUENTE_TITULO = Font(bold=True, size=14)
FUENTE_SUBTITULO = Font(size=11)
FUENTE_NOTA = Font(italic=True, size=9)
ESTILO_ENCABEZADO = Font(bold=True)
ESTILO_ENLACE = Font(color="0563C1", underline="single")

RELLENO_BLOQUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
RELLENO_AVISO = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RELLENO_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RELLENO_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RELLENO_DESCARTAT = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
BORDE_SUPERIOR = Border(top=Side(style="thin"))
AJUSTE_TEXTO = Alignment(wrap_text=True, vertical="top")

FECHA_GENERACION = datetime.now()

# Traduccion de los motivos de validar.py -- son textos fijos con datos
# incrustados (numeros, nombres de archivo), asi que se traduce por
# sustitucion de las palabras fijas, dejando el resto intacto. validar.py
# y los campos JSON no se tocan (regla de hierro del piso 8).
TRADUCCIONES_MOTIVO = [
    ("campo obligatorio vacío:", "camp obligatori buit:"),
    ("línea", "línia"),
    ("de IVA con campo vacío", "d'IVA amb camp buit"),
    ("pero cuota indica", "però la quota indica"),
    ("total no cuadra: bases+cuotas=", "el total no quadra: bases+quotes="),
    (", total indica ", ", el total indica "),
    ("nif_receptor no coincide: esperado", "el nif_receptor no coincideix: s'esperava"),
    (", encontrado ", ", s'ha trobat "),
    ("factura duplicada: mismo proveedor+num_factura que", "factura duplicada: mateix proveïdor+núm_factura que"),
    ("retención con cuota > 0: el llibre no tiene columna para representarla",
     "retenció amb quota > 0: el llibre no té columna per representar-la"),
    ("el cliente no aparece ni como emisor ni como receptor",
     "el client no apareix ni com a emissor ni com a receptor"),
    ("las dos partes son el cliente", "les dues parts són el client"),
    ("IVA incluido sin desglosar (tipo impreso:", "IVA inclòs sense desglosar (tipus imprès:"),
    ("ninguno)", "cap)"),
    ("marcada como exenta pero tiene tipo/cuota distintos de 0 -- corrige la línea (tipo 0, cuota 0) o desmarca exenta",
     "marcada com a exempta però té tipus/quota diferents de 0 -- corregeix la línia (tipus 0, quota 0) o desmarca exempta"),
]


def traducir_motivo(motivo):
    for es, ca in TRADUCCIONES_MOTIVO:
        motivo = motivo.replace(es, ca)
    return motivo


TABLA_LETRA_DNI = "TRWAGMYFPDXBNJZSQVHLCKE"


def validar_nif(nif):
    """Verificacion pura de la letra de control de un NIF espanyol (DNI,
    NIE o CIF). True/False si reconoce el formato y puede comprobarlo;
    None si no tiene forma de DNI/NIE/CIF (ej. un NIF extranjero) --
    en ese caso no se avisa, porque no sabemos si esta mal, solo que
    no lo reconocemos. No cambia ningun estado ni suma, es solo un
    aviso aparte."""
    if nif is None:
        return None
    n = "".join(c for c in nif if c.isalnum()).upper()

    if len(n) == 9 and n[:8].isdigit() and n[8].isalpha():
        return TABLA_LETRA_DNI[int(n[:8]) % 23] == n[8]

    if len(n) == 9 and n[0] in "XYZ" and n[1:8].isdigit() and n[8].isalpha():
        prefijo = {"X": "0", "Y": "1", "Z": "2"}[n[0]]
        return TABLA_LETRA_DNI[int(prefijo + n[1:8]) % 23] == n[8]

    if len(n) == 9 and n[0] in "ABCDEFGHJKLMNPQRSUVW" and n[1:8].isdigit() and (n[8].isdigit() or n[8].isalpha()):
        digitos = n[1:8]
        suma_par = sum(int(d) for d in digitos[1::2])
        suma_impar = 0
        for d in digitos[0::2]:
            doble = int(d) * 2
            suma_impar += doble // 10 + doble % 10
        digito_control = (10 - (suma_par + suma_impar) % 10) % 10
        letra_control = "JABCDEFGHI"[digito_control]
        if n[0] in "ABEH":
            return n[8] == str(digito_control)
        if n[0] in "KPQS":
            return n[8] == letra_control
        return n[8] == str(digito_control) or n[8] == letra_control

    return None


def verificar_retencion(datos):
    """True si la factura no tiene retencion, o si retencion_cuota
    cuadra con Sigma base x retencion_pct/100 (tolerancia 0,02, misma
    que el resto de la red). No cambia ningun estado ni suma."""
    retencion_pct = datos.get("retencion_pct") or 0
    retencion_cuota = datos.get("retencion_cuota") or 0
    if not retencion_pct and not retencion_cuota:
        return True
    suma_base = sum((l.get("base") or 0) for l in (datos.get("lineas_iva") or []))
    esperado = suma_base * retencion_pct / 100
    return abs(esperado - retencion_cuota) <= 0.02


def uri_fitxer(ruta):
    """Piso 13G: Excel 2007 exigeix que el Target d'un hyperlink extern
    sigui una URI valida (RFC 3986) -- confirmat amb evidencia real
    (descomprimint un .xlsx generat) que una ruta crua amb espais/
    comes fa que Excel 2007 marqui el paquet sencer com a corrupte i
    demani "reparar". file:/// + barres normals + percent-encoding."""
    ruta_resuelta = str(Path(ruta).resolve()).replace("\\", "/")
    if not ruta_resuelta.startswith("/"):
        ruta_resuelta = "/" + ruta_resuelta  # Windows: "C:/..." -> "/C:/..."
    return "file://" + quote(ruta_resuelta, safe="/:")


def ruta_desde_uri(uri):
    """Piso 13G: inversa de uri_fitxer -- nomes per a verificar_enlaces_excel,
    mai per escriure. url2pathname ja resol be la peculiaritat de
    Windows (file:///C:/... porta una barra de mes davant de la lletra
    d'unitat)."""
    return url2pathname(urlparse(uri).path)


# Piso 13G: Pla B preparat pero desactivat -- si Excel 2007 seguis
# protestant fins i tot amb una URI valida, canviar nomes aquesta
# constant a False deixa les celdes com a text pla sense hyperlink
# (els enllaços seguirien disponibles a l'informe HTML, que no es
# toca en aquest pis). Mai s'activa sol.
GENERAR_HIPERVINCLES = True


def asignar_hipervinculo(celda, ruta):
    """Piso 13G: centraliza els 7 sitis que abans feien
    `celda.hyperlink = ...; celda.font = ESTILO_ENLACE` (Piso 13E) --
    i el punt unic on el Pla B (GENERAR_HIPERVINCLES=False) es
    commutaria."""
    if GENERAR_HIPERVINCLES:
        celda.hyperlink = uri_fitxer(ruta)
        celda.font = ESTILO_ENLACE


def leer_clientes():
    with open("clientes/clientes.csv", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def cargar_validadas(carpeta):
    """Lee todos los .json de una carpeta de validadas. Si la carpeta
    no existe (el cliente no tiene ese flujo todavia), devuelve vacio."""
    facturas = []
    if not os.path.isdir(carpeta):
        return facturas
    for nombre in sorted(os.listdir(carpeta)):
        if not nombre.lower().endswith(".json"):
            continue
        with open(os.path.join(carpeta, nombre), encoding="utf-8") as f:
            facturas.append((nombre, json.load(f)))
    return facturas


def cargar_decisiones(carpeta_cliente):
    """Lee decisions.csv (archivo,accion,nota,qui,data) si existe. Sin
    archivo, o vacio, devuelve {}. Solo aplica a facturas ya validadas
    (los ERROR no tienen ficha, no son aprobables).

    Piso 13M: decisions.csv es llibre major (app.py ja no sobreescriu
    cap fila, sempre n'afegeix una) -- l'estat efectiu d'un archiu es
    la seva ULTIMA fila; si es "revertir", es treu del diccionari
    (com si mai s'hagues decidit). Res mes en aquest fitxer necessita
    canviar: sumar_bloque/_escribir_fila_detalle ja fan
    decisiones.get(nombre) i decision.get("accion") == "aprovar"."""
    ruta = f"{carpeta_cliente}/decisions.csv"
    decisiones = {}
    if not os.path.exists(ruta):
        return decisiones
    with open(ruta, encoding="utf-8") as f:
        for fila in csv.DictReader(f):
            archivo = fila.get("archivo")
            if not archivo:
                continue
            if fila.get("accion") == "revertir":
                decisiones.pop(archivo, None)
            else:
                decisiones[archivo] = fila
    return decisiones


def estat_efectiu(decision):
    """Piso 13Y: 'aprovar', 'descartar', o None -- la interpretacio
    d'una decisio ja obtinguda de cargar_decisiones().get(archivo). Es
    exactament l'expressio que es repetia a sumar_bloque,
    escribir_detalle i _escribir_fila_detalle -- cap canvi de gramatica,
    nomes un sol lloc si mai cal canviar-la."""
    return decision.get("accion") if decision else None


def historial_decisiones(carpeta_cliente, archivo):
    """Piso 13M: igual que en app.py -- totes les files d'aquest
    archiu a decisions.csv, en ordre cronologic."""
    ruta = f"{carpeta_cliente}/decisions.csv"
    if not os.path.exists(ruta):
        return []
    with open(ruta, encoding="utf-8") as f:
        return [fila for fila in csv.DictReader(f) if fila.get("archivo") == archivo]


TRADUCCION_ACCION = {"aprovar": "aprovada", "descartar": "descartada", "revertir": "revertit"}


def resumen_historial_decisiones(historial):
    """Piso 13M: igual que en app.py -- text compacte, None si no hi
    ha cap "revertir" al historial (sense soroll per a la resta)."""
    if not any(fila.get("accion") == "revertir" for fila in historial):
        return None
    return " — ".join(
        f"{TRADUCCION_ACCION.get(fila.get('accion'), fila.get('accion'))} per {fila.get('qui')} el {fila.get('data')}"
        for fila in historial
    )


def trimestre_de(fecha):
    """'2026-04-21' -> '2T'. Si la fecha falta o no se puede leer, None."""
    if not fecha:
        return None
    try:
        mes = int(fecha[5:7])
    except (ValueError, IndexError):
        return None
    return f"{(mes - 1) // 3 + 1}T"


def encontrar_original(carpeta_origen, nombre_json):
    """Busca el original directamente en carpeta_origen, y si no esta
    ahi, en cualquier subcarpeta hermana no reservada del pipeline --
    algunos clientes (davinstal) organizan sus facturas de compra por
    proveedor (rebudes/biosca/, rebudes/SALTOKI/...) en vez de dejarlas
    sueltas en rebudes/entrada/. Piso 9.3: "procesadas" SI se busca
    aqui (a diferencia de listar_archivos_rebudes) -- un original ya
    movido ahi tras procesarse sigue siendo su ubicacion legitima."""
    base = os.path.splitext(nombre_json)[0]
    for ext in EXTENSIONES_ORIGINAL:
        ruta = os.path.join(carpeta_origen, base + ext)
        if os.path.exists(ruta):
            return ruta
    if os.path.isdir(carpeta_origen):
        for nombre_sub in sorted(os.listdir(carpeta_origen)):
            ruta_sub = os.path.join(carpeta_origen, nombre_sub)
            if not os.path.isdir(ruta_sub) or nombre_sub.lower() in SUBCARPETAS_NO_ORIGINALES:
                continue
            for ext in EXTENSIONES_ORIGINAL:
                ruta = os.path.join(ruta_sub, base + ext)
                if os.path.exists(ruta):
                    return ruta
    return None


def listar_archivos_rebudes(carpeta_rebudes):
    """Igual que en extraer_todas.py/informe.py: cuenta entrada/ y
    cualquier subcarpeta hermana no reservada (proveedores organizados
    por subcarpeta, ej. davinstal). Solo lectura -- para detectar
    errores (presente sin ficha)."""
    rutas = []
    if not os.path.isdir(carpeta_rebudes):
        return rutas
    for nombre in sorted(os.listdir(carpeta_rebudes)):
        ruta = os.path.join(carpeta_rebudes, nombre)
        if not os.path.isdir(ruta) or nombre.lower() in SUBCARPETAS_RESERVADAS:
            continue
        for nombre_archivo in sorted(os.listdir(ruta)):
            if nombre_archivo.lower().endswith(EXTENSIONES_ORIGINAL):
                rutas.append(os.path.join(ruta, nombre_archivo))
    return rutas


def listar_archivos_planos(carpeta):
    """Listado plano (sin subcarpetas) -- para el origen de ingressos,
    que no se organiza por proveedor."""
    if not os.path.isdir(carpeta):
        return []
    return [
        os.path.join(carpeta, f) for f in sorted(os.listdir(carpeta))
        if f.lower().endswith(EXTENSIONES_ORIGINAL)
    ]


def detectar_errores(rutas_presentes, carpeta_extraidas):
    """Archivos presentes sin ficha extraida -- unica senal disponible,
    porque extraer_todas.py no persiste el motivo de un fallo de
    extraccion. Solo lectura de disco."""
    extraidos = set()
    if os.path.isdir(carpeta_extraidas):
        extraidos = {
            os.path.splitext(f)[0] for f in os.listdir(carpeta_extraidas) if f.lower().endswith(".json")
        }
    return [r for r in rutas_presentes if os.path.splitext(os.path.basename(r))[0] not in extraidos]


MOTIVO_ERROR_GENERICO = (
    "No s'ha pogut generar la fitxa — l'arxiu és present però no hi ha extracció. "
    "Cal revisar l'escaneig o tornar-ho a intentar."
)

MOTIVO_ERROR_CORRUPTE = (
    "L'arxiu original és present però buit o corromput — probablement un "
    "problema de sincronització. Cal tornar-lo a sincronitzar o demanar-lo de nou."
)


def archivo_corrupto(ruta):
    """Igual que en informe.py: True si la ruta existe pero esta vacia,
    o es un PDF sense trailer %%EOF en l'ultim KB (truncat a mitges --
    p. ex. per una sincronitzacio d'iCloud interrompuda). Piso 9.3."""
    if not os.path.exists(ruta):
        return False
    tamano = os.path.getsize(ruta)
    if tamano == 0:
        return True
    if ruta.lower().endswith(".pdf"):
        with open(ruta, "rb") as f:
            f.seek(max(0, tamano - 1024))
            cola = f.read()
        if b"%%EOF" not in cola:
            return True
    return False


def motivo_error(ruta):
    """Motiu derivat -- no inventat -- per a un archiu ERROR: distingeix
    corrupcio comprovable en disc del generic 'no hi ha extraccio'."""
    return MOTIVO_ERROR_CORRUPTE if archivo_corrupto(ruta) else MOTIVO_ERROR_GENERICO


def verificar_enlaces_excel(ruta_xlsx, carpeta_cliente):
    """Piso 9.3: reabre el Excel ya escrito y comprueba cada hyperlink.
    'Trencat' (no existe) es un bug de codigo y no deberia pasar nunca;
    'corrupte' (existe pero vacio o truncado) es un problema de datos
    que ningun cambio de codigo puede arreglar, solo informar.

    Piso 13G: el target ya es siempre una URI file:// absoluta (nunca
    relativa a carpeta_cliente) -- se convierte con ruta_desde_uri en
    vez de unirla con os.path.join."""
    wb = load_workbook(ruta_xlsx)
    verificados = 0
    trencats = []
    corruptes = []
    manuals_sense_enllac = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            # Piso 14: una fila d'ENTRADA MANUAL mai té original -- es
            # compta a part, informatiu, MAI com a trencat (ja no arriba
            # a comptar-se com a candidat: cell.hyperlink es sempre None).
            if any(isinstance(c.value, str) and "ENTRADA MANUAL" in c.value for c in row):
                manuals_sense_enllac += 1
            for cell in row:
                if cell.hyperlink is None:
                    continue
                verificados += 1
                ruta = ruta_desde_uri(cell.hyperlink.target)
                if not os.path.exists(ruta):
                    trencats.append(cell.hyperlink.target)
                elif archivo_corrupto(ruta):
                    corruptes.append(cell.hyperlink.target)
    print(
        f"{os.path.basename(carpeta_cliente)} / excel: {verificados} enllaços verificats, "
        f"{len(trencats)} trencats, {len(corruptes)} corruptes, "
        f"{manuals_sense_enllac} manual sense enllaç"
    )
    for href in trencats:
        print(f"  TRENCAT: {href}")
    return trencats, corruptes


PATRON_SUM = re.compile(r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$")
PATRON_RESTA_SUM = re.compile(r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)-SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$")
# Piso 14: =SUM() amb una llista de cel·les concretes en comptes d'un
# rang contigu -- veure _formula_suma_filas i el docstring nou de
# escribir_detalle (les línies d'un cub ja no cauen sempre seguides).
PATRON_SUM_MULTI = re.compile(r"^=SUM\(([A-Z]+\d+(?:,[A-Z]+\d+)*)\)$")


def _leer_rango(ws, columna, ini, fin):
    total = 0.0
    for f in range(ini, fin + 1):
        valor = ws[f"{columna}{f}"].value
        total += valor or 0
    return total


def _leer_filas(ws, columna, filas):
    """Piso 14: mateixa suma que _leer_rango pero per a una llista de
    files concretes (no un rang contigu) -- veure PATRON_SUM_MULTI."""
    total = 0.0
    for f in filas:
        valor = ws[f"{columna}{f}"].value
        total += valor or 0
    return total


def _mapa_hojas_xml(zf):
    """Piso 13J: averigua a que xl/worksheets/sheetN.xml corresponde
    cada nombre de hoja visible ("2T", "AVISOS"...). No se puede asumir
    el orden (sheet1.xml = primera hoja creada): hay que leerlo de
    verdad de xl/workbook.xml (nombre -> r:id) y xl/_rels/workbook.xml.rels
    (r:id -> ruta del archivo), que es como el propio Excel lo resuelve.
    Los atributos de cada etiqueta no siempre vienen en el mismo orden
    (comprobado en un archivo real), por eso se extraen uno a uno en vez
    de asumir una posicion fija."""
    workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
    rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    rid_a_target = {}
    for etiqueta in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
        id_m = re.search(r'Id="([^"]+)"', etiqueta)
        target_m = re.search(r'Target="([^"]+)"', etiqueta)
        if id_m and target_m:
            rid_a_target[id_m.group(1)] = target_m.group(1)

    mapa = {}
    for etiqueta in re.findall(r"<sheet\b[^>]*/>", workbook_xml):
        nombre_m = re.search(r'name="([^"]+)"', etiqueta)
        rid_m = re.search(r'r:id="([^"]+)"', etiqueta)
        if not (nombre_m and rid_m):
            continue
        target = rid_a_target[rid_m.group(1)]
        # Target puede venir absoluto ("/xl/worksheets/sheet1.xml") o
        # relativo a xl/ ("worksheets/sheet1.xml") -- visto ambos casos
        # en el mismo archivo real (rels de hojas vs. de estilos/tema).
        ruta = target[1:] if target.startswith("/") else f"xl/{target}"
        mapa[nombre_m.group(1)] = ruta
    return mapa


def guardar_amb_reintents(fn_escriure):
    """Piso 13W: abans, un sol PermissionError es donava per "Excel
    obert" i es rendia a la primera -- pero a Windows, PermissionError
    (WinError 5/32) es la MATEIXA excepcio tant si Excel el te obert
    de veritat com si es un antivirus, OneDrive/iCloud sincronitzant,
    o l'atribut de nomes-lectura: no hi ha manera de distingir la
    causa des del tipus d'excepcio. En comptes d'endevinar, es
    reintenta -- la majoria de bloquejos son transitoris i es
    resolen sols en segons. Intent 1 immediat; si falla, 5 reintents
    mes amb esperes creixents 0.5/1/2/3/4s (~10.5s totals, 6 intents
    en total). Retorna (True, None) si reeix, (False, ultim_error) si
    els 6 fallen -- el crider tria el missatge exacte (cada fitxer te
    el seu propi nom i verb)."""
    esperes = [0, 0.5, 1, 2, 3, 4]
    ultim_error = None
    for espera in esperes:
        if espera:
            time.sleep(espera)
        try:
            fn_escriure()
            return True, None
        except PermissionError as e:
            ultim_error = e
    return False, ultim_error


def inyectar_valores_cacheados(ruta_excel, comprobaciones_por_hoja):
    """Piso 13J: arregla la causa encontrada en el Piso 13I (dissecio
    forense) de que Excel 2007 pida "reparar" un archivo recien
    generado. openpyxl escribe SIEMPRE toda celda de formula como
    <f>...</f><v /> (v vacio) -- confirmado leyendo su propio codigo
    fuente (_writer.py) y con una reproduccion minima aislada. El
    espec ECMA-376 permite omitir <v> del todo (minOccurs="0"), pero
    dejarlo vacio en una celda que por defecto se declara numerica
    (el atributo t no se escribe) es lo que Excel 2007 rechaza al
    cargar -- Excel moderno lo tolera en silencio y recalcula.

    El propio codigo YA sabe el valor exacto de cada formula antes de
    escribirla (comprobaciones_por_hoja, la misma lista que usa
    verificar_formulas para comparar al centimo) -- este paso solo
    inyecta ese numero ya conocido dentro de <v>, reabriendo el .xlsx
    (un zip) recien guardado. openpyxl no expone ninguna forma de
    hacerlo por su API publica, asi que se hace directamente sobre el
    XML ya escrito, y se reescribe el archivo entero (los zip no se
    editan por partes)."""
    with zipfile.ZipFile(ruta_excel) as zf:
        contenidos = {nombre: zf.read(nombre) for nombre in zf.namelist()}
        mapa_hojas = _mapa_hojas_xml(zf)

    for nombre_hoja, comprobaciones in comprobaciones_por_hoja.items():
        ruta_hoja = mapa_hojas[nombre_hoja]
        xml = contenidos[ruta_hoja].decode("utf-8")
        for comprobacion in comprobaciones:
            coordenada, valor = comprobacion[0], comprobacion[-1]
            if valor == 0:
                valor = 0  # evita "-0.0" si una resta redondea a cero negativo
            patron = re.compile(r'(<c r="' + re.escape(coordenada) + r'"[^>]*><f>[^<]*</f>)<v\s*/?>(?:</v>)?')
            xml, n = patron.subn(lambda m: f"{m.group(1)}<v>{valor}</v>", xml, count=1)
            if n != 1:
                raise ValueError(f"{nombre_hoja}!{coordenada}: no s'ha trobat la cel·la de fórmula per injectar el valor")
        contenidos[ruta_hoja] = xml.encode("utf-8")

    ruta_temp = ruta_excel + ".tmp"
    with zipfile.ZipFile(ruta_temp, "w", zipfile.ZIP_DEFLATED) as zf_out:
        for nombre, datos in contenidos.items():
            zf_out.writestr(nombre, datos)
    os.replace(ruta_temp, ruta_excel)


def verificar_formulas(ruta_xlsx, comprobaciones_por_hoja):
    """Piso 12A: openpyxl no calcula formulas -- solo Excel, al abrir el
    archivo. Por eso, tras guardar, se reabre el mismo Excel (load_workbook
    devuelve el TEXTO de la formula, nunca su resultado) y por cada
    formula escrita: (1) se reparsa su rango con una regex; (2) se
    comprueba que coincide EXACTO con el rango anotado al escribir
    DETALL (para detectar un futuro cambio que desalinee los bloques);
    (3) se suman en Python los valores planos de esas celdas y se
    compara contra el valor que sumar_bloque ya calculo en memoria
    durante este mismo run, con tolerancia de centimo. Cualquier fallo
    para el run con un error claro -- regla 4, nada en silencio."""
    wb = load_workbook(ruta_xlsx)
    total_verificadas = 0
    for nombre_hoja, comprobaciones in comprobaciones_por_hoja.items():
        ws = wb[nombre_hoja]
        for comprobacion in comprobaciones:
            coordenada = comprobacion[0]
            formula = ws[coordenada].value

            if comprobacion[1] == "resta":
                _, _, (col_i, ini_i, fin_i), (col_g, ini_g, fin_g), valor_esperado = comprobacion
                m = PATRON_RESTA_SUM.match(str(formula))
                if not m:
                    raise ValueError(f"{nombre_hoja}!{coordenada}: fórmula inesperada: {formula!r}")
                col_p1, ini_p1, col_p1b, fin_p1, col_p2, ini_p2, col_p2b, fin_p2 = m.groups()
                rango_parseado = (col_p1, int(ini_p1), col_p1b, int(fin_p1), col_p2, int(ini_p2), col_p2b, int(fin_p2))
                rango_esperado = (col_i, ini_i, col_i, fin_i, col_g, ini_g, col_g, fin_g)
                if rango_parseado != rango_esperado:
                    raise ValueError(
                        f"{nombre_hoja}!{coordenada}: rang de la fórmula ({rango_parseado}) "
                        f"no coincideix amb l'anotat ({rango_esperado})"
                    )
                valor_real = _leer_rango(ws, col_i, ini_i, fin_i) - _leer_rango(ws, col_g, ini_g, fin_g)
            elif comprobacion[1] == "multi":
                _, _, columna, filas, valor_esperado = comprobacion
                m = PATRON_SUM_MULTI.match(str(formula))
                if not m:
                    raise ValueError(f"{nombre_hoja}!{coordenada}: fórmula inesperada: {formula!r}")
                celdas_parseadas = m.group(1).split(",")
                celdas_esperadas = [f"{columna}{f}" for f in filas]
                if celdas_parseadas != celdas_esperadas:
                    raise ValueError(
                        f"{nombre_hoja}!{coordenada}: cel·les de la fórmula ({celdas_parseadas}) "
                        f"no coincideixen amb les anotades ({celdas_esperadas})"
                    )
                valor_real = _leer_filas(ws, columna, filas)
            else:
                _, columna, ini, fin, valor_esperado = comprobacion
                m = PATRON_SUM.match(str(formula))
                if not m:
                    raise ValueError(f"{nombre_hoja}!{coordenada}: fórmula inesperada: {formula!r}")
                col_p1, ini_p, col_p2, fin_p = m.groups()
                if (col_p1, int(ini_p), col_p2, int(fin_p)) != (columna, ini, columna, fin):
                    raise ValueError(
                        f"{nombre_hoja}!{coordenada}: rang de la fórmula ({col_p1}{ini_p}:{col_p2}{fin_p}) "
                        f"no coincideix amb l'anotat ({columna}{ini}:{columna}{fin})"
                    )
                valor_real = _leer_rango(ws, columna, ini, fin)

            if abs(valor_real - valor_esperado) > 0.01:
                raise ValueError(
                    f"{nombre_hoja}!{coordenada}: la fórmula suma {valor_real:.2f}, "
                    f"s'esperava {valor_esperado:.2f}"
                )
            total_verificadas += 1

    return total_verificadas


def escribir_titulo(ws, nombre_cliente, nif_cliente):
    ws.cell(row=1, column=1, value=f"SUMATORIS {nombre_cliente}").font = FUENTE_TITULO
    ws.cell(row=2, column=1, value=f"NIF {nif_cliente}").font = FUENTE_SUBTITULO
    ws.cell(row=3, column=1, value="Ejercici 2026").font = FUENTE_SUBTITULO
    # Piso 13O: objecte datetime real (no text) -- mateixa tecnica que
    # FORMATO_MONEDA (text literal dins del number_format, Piso 13F),
    # aixi la cel·la es veu identica ("Generat el 14/07/2026 10:30")
    # pero el valor subjacent ja es una data de veritat.
    celda_generat = ws.cell(row=4, column=1, value=FECHA_GENERACION)
    celda_generat.number_format = '"Generat el "dd/mm/yyyy hh:mm'
    celda_generat.font = FUENTE_SUBTITULO
    return 6  # fila 5 en blanco, el resto empieza en la 6


def configurar_impresion(ws):
    """Piso 13E: nomes presentacio -- prepara el full per imprimir-se
    be en paper real (horitzontal, ajustat a 1 pagina d'ample, la
    capçalera de escribir_titulo repetida a cada pagina, marges
    estrets). Cap dada ni logica canvia."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def sumar_bloque(facturas, decisiones):
    """Suma un bloque (gastos o ingresos) de un trimestre: base/cuota
    por tipo de IVA, total y retencion de las que cuentan, y aparte
    las que hay que revisar y las descartadas.

    decisiones (de decisions.csv) puede anular el estado de una ficha:
    'aprovar' hace que una REVISAR cuente como si fuera OK; 'descartar'
    hace que nunca cuente (ni las OK), y va a su propio bloque. Sin
    decisiones.csv, o vacio, el comportamiento es identico al de
    siempre (ninguna entrada tiene decision).

    Piso 13R: las fichas exenta=true forman su propio cubo "exenta"
    -- separado del cubo del tipo 0% real, que sigue significando
    exactamente lo mismo que antes. La cuota de una exenta coherente
    es siempre 0 (lo garantiza validar.py antes de OK); Piso 14C:
    una línia exempta APROVADA manualment tot i ser incoherent pot
    portar quota real distinta de 0 -- ja no s'amaga (es suma i
    s'escriu igual que qualsevol altre tipus).

    Piso 14B: exenta es un campo POR LÍNIA (validar.py ja el resol i
    deriva el de factura) -- una factura mixta (una línia exempta,
    una altra al 21%) reparteix cada línia al seu cub."""
    sumas = {tipo: {"base": 0.0, "cuota": 0.0} for tipo in TIPOS_IVA}
    sumas["otros"] = {"base": 0.0, "cuota": 0.0}
    sumas["exenta"] = {"base": 0.0, "cuota": 0.0}
    total_ok = 0.0
    retencion_ok = 0.0
    revisar = []
    descartados = []

    for nombre, datos in facturas:
        decision = decisiones.get(nombre)

        if estat_efectiu(decision) == "descartar":
            descartados.append((nombre, datos, decision))
            continue

        cuenta = datos.get("estado") == "OK" or estat_efectiu(decision) == "aprovar"
        if not cuenta:
            revisar.append((nombre, datos))
            continue

        for linea in datos.get("lineas_iva") or []:
            tipo = linea.get("tipo_iva")
            if linea.get("exenta"):
                clave = "exenta"
            else:
                clave = tipo if tipo in TIPOS_IVA else "otros"
            sumas[clave]["base"] += linea.get("base") or 0
            sumas[clave]["cuota"] += linea.get("cuota") or 0
        total_ok += datos.get("total") or 0
        retencion_ok += datos.get("retencion_cuota") or 0

    return sumas, total_ok, retencion_ok, revisar, descartados


def escribir_bloque(ws, fila, titulo, sumas, con_retencion, etiqueta_retencion="TOTAL RETENCIONS"):
    """Piso 12A: escribe solo la ESTRUCTURA (etiquetas, tipos de IVA
    presentes) -- las celdas de Base/Quota/Total/TOTAL RETENCIONS quedan en
    blanco, se rellenan mas tarde con formulas en escribir_formulas_bloque
    una vez se conoce el rango de filas de DETALL. Devuelve (fila,
    info) -- info trae la posicion de cada celda que hay que rellenar.

    Piso 13S: etiqueta_retencion permet que DESPESES digui "Σ RETENCIONS
    SUPORTADES" (les del lloguer, model 115) sense tocar el text ja
    establert d'INGRESSOS ("TOTAL RETENCIONS")."""
    for col in range(1, 5):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value=titulo)
    fila += 1

    for col, texto in enumerate(["Tipus IVA", "Base", "Quota", "Total"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    tipos_a_escribir = list(TIPOS_IVA)
    if sumas["otros"]["base"] or sumas["otros"]["cuota"]:
        tipos_a_escribir.append("otros")
    # Piso 13R: cub propi "exenta" -- nomes apareix si hi ha alguna
    # base exempta a sumar, mateix criteri que "otros".
    if sumas["exenta"]["base"] or sumas["exenta"]["cuota"]:
        tipos_a_escribir.append("exenta")

    filas_por_tipo = {}
    for tipo in tipos_a_escribir:
        etiqueta_tipo = "ALTRES" if tipo == "otros" else ("EXEMPTA" if tipo == "exenta" else tipo)
        ws.cell(row=fila, column=1, value=etiqueta_tipo)
        ws.cell(row=fila, column=2).number_format = FORMATO_MONEDA
        ws.cell(row=fila, column=3).number_format = FORMATO_MONEDA
        filas_por_tipo[tipo] = fila
        fila += 1

    for col in range(1, 5):
        ws.cell(row=fila, column=col).border = BORDE_SUPERIOR
    ws.cell(row=fila, column=1, value="TOTAL").font = ESTILO_ENCABEZADO
    for col in (2, 3, 4):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.number_format = FORMATO_MONEDA
    fila_total = fila
    fila += 1

    fila_retencio = None
    fila_total_fra = None
    if con_retencion:
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = BORDE_SUPERIOR
        ws.cell(row=fila, column=1, value=etiqueta_retencion).font = ESTILO_ENCABEZADO
        celda = ws.cell(row=fila, column=4)
        celda.font = ESTILO_ENCABEZADO
        celda.number_format = FORMATO_MONEDA
        fila_retencio = fila
        fila += 1

        # Piso 14C: "Σ TOTAL FRA." AL COSTAT de "TOTAL" (mai en lloc
        # seu) -- el seu llibre defineix TOTAL FRA. = base+IVA-retenció,
        # les dues xifres (bruta i neta) es queden sempre visibles.
        ws.cell(row=fila, column=1, value="Σ TOTAL FRA.").font = ESTILO_ENCABEZADO
        celda_total_fra = ws.cell(row=fila, column=4)
        celda_total_fra.font = ESTILO_ENCABEZADO
        celda_total_fra.number_format = FORMATO_MONEDA
        fila_total_fra = fila
        fila += 1

    info = {
        "filas_por_tipo": filas_por_tipo, "fila_total": fila_total,
        "fila_retencio": fila_retencio, "fila_total_fra": fila_total_fra,
    }
    return fila + 1, info  # una fila en blanco antes del siguiente bloque


def _formula_suma_filas(columna, filas):
    """Piso 14: =SUM() amb una llista de cel·les concretes (no un rang
    contigu) -- sintaxi Excel vàlida (comes separen referències dins
    d'un mateix SUM). Necessari des que escribir_detalle ja no ordena
    per tipus primer: les línies d'un cub poden caure disperses entre
    factures (cada una contigua per si mateixa, mai les unes amb les
    altres)."""
    return f"=SUM({','.join(f'{columna}{f}' for f in filas)})"


def escribir_formulas_bloque(ws, info, rango_detalle, sumas, total_ok, retencion_ok):
    """Piso 12A: rellena con formulas =SUM() las celdas que escribir_bloque
    dejo en blanco, apuntando SOLO al rango de "FACTURES QUE SUMEN" de
    DETALL (rango_detalle, devuelto por escribir_detalle). Si un tipo no
    tiene ninguna linia que sume, se escribe 0 (no hay rango que sumar,
    no tiene sentido una formula). Devuelve la lista de comprobaciones
    para verificar_formulas.

    Piso 14: rango_detalle["tipos"][tipo] es ahora una LISTA de filas
    exactas (pueden estar dispersas entre facturas, ya no un (ini, fin)
    contiguo) -- comprobaciones para estas celdas llevan el
    discriminador "multi" (filas, no ini/fin) en vez del formato viejo.
    TOTAL/retención siguen sobre rango_detalle["bloque"], que SIGUE
    siendo un rango contiguo (el bloque entero nunca se fragmenta,
    solo los cubos por tipo dentro de el)."""
    comprobaciones = []

    for tipo, fila_tipo in info["filas_por_tipo"].items():
        filas_tipo = rango_detalle["tipos"].get(tipo)
        celda_base = ws.cell(row=fila_tipo, column=2)
        celda_cuota = ws.cell(row=fila_tipo, column=3)
        if not filas_tipo:
            celda_base.value = 0
            celda_cuota.value = 0
            continue
        celda_base.value = _formula_suma_filas(COLUMNA_BASE_DETALLE, filas_tipo)
        comprobaciones.append((celda_base.coordinate, "multi", COLUMNA_BASE_DETALLE, tuple(filas_tipo), round(sumas[tipo]["base"], 2)))
        # Piso 14C: la quota Exempta ja no es queda buida per
        # construcció -- una línia exempta APROVADA manualment tot i
        # ser incoherent (tipus/quota diferents de 0) pot portar una
        # quota real distinta de 0, i s'ha de veure (RESULTAT, a
        # escribir_formulas_resultat, ja la sumava be des de sempre;
        # nomes aquest cub la deixava invisible).
        celda_cuota.value = _formula_suma_filas(COLUMNA_QUOTA_DETALLE, filas_tipo)
        comprobaciones.append((celda_cuota.coordinate, "multi", COLUMNA_QUOTA_DETALLE, tuple(filas_tipo), round(sumas[tipo]["cuota"], 2)))

    rango_bloque = rango_detalle["bloque"]
    fila_total = info["fila_total"]
    celda_total_base = ws.cell(row=fila_total, column=2)
    celda_total_cuota = ws.cell(row=fila_total, column=3)
    celda_total_total = ws.cell(row=fila_total, column=4)
    if rango_bloque is None:
        celda_total_base.value = 0
        celda_total_cuota.value = 0
        celda_total_total.value = 0
    else:
        ini, fin = rango_bloque
        suma_base_esperada = round(sum(sumas[t]["base"] for t in sumas), 2)
        suma_cuota_esperada = round(sum(sumas[t]["cuota"] for t in sumas), 2)
        celda_total_base.value = f"=SUM({COLUMNA_BASE_DETALLE}{ini}:{COLUMNA_BASE_DETALLE}{fin})"
        celda_total_cuota.value = f"=SUM({COLUMNA_QUOTA_DETALLE}{ini}:{COLUMNA_QUOTA_DETALLE}{fin})"
        celda_total_total.value = f"=SUM({COLUMNA_TOTAL_DETALLE}{ini}:{COLUMNA_TOTAL_DETALLE}{fin})"
        comprobaciones.append((celda_total_base.coordinate, COLUMNA_BASE_DETALLE, ini, fin, suma_base_esperada))
        comprobaciones.append((celda_total_cuota.coordinate, COLUMNA_QUOTA_DETALLE, ini, fin, suma_cuota_esperada))
        comprobaciones.append((celda_total_total.coordinate, COLUMNA_TOTAL_DETALLE, ini, fin, round(total_ok, 2)))

    if info["fila_retencio"] is not None:
        celda_retencio = ws.cell(row=info["fila_retencio"], column=4)
        if rango_bloque is None:
            celda_retencio.value = 0
        else:
            ini, fin = rango_bloque
            celda_retencio.value = f"=SUM({COLUMNA_RETENCIO_DETALLE}{ini}:{COLUMNA_RETENCIO_DETALLE}{fin})"
            comprobaciones.append((celda_retencio.coordinate, COLUMNA_RETENCIO_DETALLE, ini, fin, round(retencion_ok, 2)))

    # Piso 14C: "Σ TOTAL FRA." -- mateix patró exacte que TOTAL/
    # retenció (rang contigu sobre rango_bloque, mai fragmentat), sumant
    # la columna TOTAL FRA. (cada factura nomes hi escriu a la primera
    # línia, com Total/Retenció -- una fila en blanc suma 0).
    if info["fila_total_fra"] is not None:
        celda_total_fra = ws.cell(row=info["fila_total_fra"], column=4)
        if rango_bloque is None:
            celda_total_fra.value = 0
        else:
            ini, fin = rango_bloque
            celda_total_fra.value = f"=SUM({COLUMNA_LIQUID_DETALLE}{ini}:{COLUMNA_LIQUID_DETALLE}{fin})"
            comprobaciones.append((
                celda_total_fra.coordinate, COLUMNA_LIQUID_DETALLE, ini, fin, round(total_ok - retencion_ok, 2),
            ))

    return comprobaciones


def escribir_resultat_estructura(ws, fila, sumas_g, sumas_i):
    """Piso 12A: RESULTAT DEL TRIMESTRE, protagonista al inicio de la
    hoja -- escribe solo la ESTRUCTURA (etiquetas, tipos de IVA
    presentes), dejando en blanco las celdas de Quota despeses/Quota
    ingressos/RESULTAT IVA -- se rellenan en escribir_formulas_resultat
    una vez se conoce el rango de filas de DETALL, con el MISMO =SUM()
    que usan los bloques DESPESES/INGRESSOS (dos formulas
    independientes sobre el mismo rango, no una referencia a la otra
    celda -- asi cada formula del libro se puede verificar sola, sin
    encadenar una formula que lee otra formula). No es una liquidacion
    oficial, solo un resultat de treball."""
    for col in range(1, 4):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value="RESULTAT DEL TRIMESTRE")
    fila += 1

    for col, texto in enumerate(["Tipus IVA", "Quota despeses", "Quota ingressos"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    tipos_a_escribir = list(TIPOS_IVA)
    if sumas_g["otros"]["cuota"] or sumas_i["otros"]["cuota"]:
        tipos_a_escribir.append("otros")

    filas_por_tipo = {}
    for tipo in tipos_a_escribir:
        etiqueta_tipo = "ALTRES" if tipo == "otros" else tipo
        ws.cell(row=fila, column=1, value=etiqueta_tipo)
        ws.cell(row=fila, column=2).number_format = FORMATO_MONEDA
        ws.cell(row=fila, column=3).number_format = FORMATO_MONEDA
        filas_por_tipo[tipo] = fila
        fila += 1

    for col in range(1, 4):
        ws.cell(row=fila, column=col).border = BORDE_SUPERIOR
    ws.cell(row=fila, column=1, value="RESULTAT IVA (repercutit - suportat)").font = ESTILO_ENCABEZADO
    ws.cell(row=fila, column=3).font = ESTILO_ENCABEZADO
    ws.cell(row=fila, column=3).number_format = FORMATO_MONEDA
    fila_resultat = fila
    fila += 1

    nota = (
        "IVA cobrat menys IVA pagat; positiu = a ingressar, negatiu = a compensar "
        "(resultat de treball, no liquidació oficial)"
    )
    ws.cell(row=fila, column=1, value=nota).font = FUENTE_NOTA
    fila += 1

    # Piso 14C: nota de doctrina -- perquè ningú confongui TOTAL FRA.
    # (la xifra que fa servir el llibre de la gestoria) amb la base
    # del 303 (que es calcula sempre amb les quotes, mai amb aquesta
    # columna).
    nota_total_fra = (
        "TOTAL FRA. = base + IVA − IRPF (com al llibre); el 303 es calcula amb les quotes."
    )
    ws.cell(row=fila, column=1, value=nota_total_fra).font = FUENTE_NOTA
    fila += 1

    info = {"filas_por_tipo": filas_por_tipo, "fila_resultat": fila_resultat}
    return fila + 1, info


def escribir_formulas_resultat(ws, info, rango_g, rango_i, sumas_g, sumas_i):
    """Piso 12A: rellena RESULTAT DEL TRIMESTRE con formulas =SUM()
    sobre los MISMOS rangos de DETALL que ya usan DESPESES/INGRESSOS
    (rango_g/rango_i, devueltos por escribir_detalle) -- la misma
    cifra, calculada de forma independiente. RESULTAT IVA es una unica
    formula con las dos SUM() restadas (=SUM(ingressos)-SUM(despeses)),
    tambien referida directamente a DETALL, nunca a otra celda-formula."""
    comprobaciones = []

    for tipo, fila_tipo in info["filas_por_tipo"].items():
        filas_g = rango_g["tipos"].get(tipo)
        celda_g = ws.cell(row=fila_tipo, column=2)
        if not filas_g:
            celda_g.value = 0
        else:
            celda_g.value = _formula_suma_filas(COLUMNA_QUOTA_DETALLE, filas_g)
            comprobaciones.append((celda_g.coordinate, "multi", COLUMNA_QUOTA_DETALLE, tuple(filas_g), round(sumas_g[tipo]["cuota"], 2)))

        filas_i = rango_i["tipos"].get(tipo)
        celda_i = ws.cell(row=fila_tipo, column=3)
        if not filas_i:
            celda_i.value = 0
        else:
            celda_i.value = _formula_suma_filas(COLUMNA_QUOTA_DETALLE, filas_i)
            comprobaciones.append((celda_i.coordinate, "multi", COLUMNA_QUOTA_DETALLE, tuple(filas_i), round(sumas_i[tipo]["cuota"], 2)))

    fila_resultat = info["fila_resultat"]
    celda_resultat = ws.cell(row=fila_resultat, column=3)
    rango_bloque_g = rango_g["bloque"]
    rango_bloque_i = rango_i["bloque"]
    cuota_g_total = round(sum(sumas_g[t]["cuota"] for t in sumas_g), 2)
    cuota_i_total = round(sum(sumas_i[t]["cuota"] for t in sumas_i), 2)
    if rango_bloque_g is not None and rango_bloque_i is not None:
        ini_g, fin_g = rango_bloque_g
        ini_i, fin_i = rango_bloque_i
        celda_resultat.value = (
            f"=SUM({COLUMNA_QUOTA_DETALLE}{ini_i}:{COLUMNA_QUOTA_DETALLE}{fin_i})"
            f"-SUM({COLUMNA_QUOTA_DETALLE}{ini_g}:{COLUMNA_QUOTA_DETALLE}{fin_g})"
        )
        comprobaciones.append((
            celda_resultat.coordinate, "resta",
            (COLUMNA_QUOTA_DETALLE, ini_i, fin_i), (COLUMNA_QUOTA_DETALLE, ini_g, fin_g),
            round(cuota_i_total - cuota_g_total, 2),
        ))
    else:
        celda_resultat.value = round(cuota_i_total - cuota_g_total, 2)

    return comprobaciones


# Piso 13S: "% Ret." nova, "Retenció" renombrada "Retenció (€)" per
# claredat. Totes les fórmules =SUM() referencien NOMÉS via aquestes
# constants, mai un literal cru -- canvi mecànic.
# Piso 13Y: Líquid (avui "TOTAL FRA.") es NOMES informatiu (total -
# retencio, o el total si no n'hi ha) -- els cubs per tipus d'IVA, el
# RESULTAT i tota la resta de matematica fiscal segueixen SEMPRE sobre
# base+quota, mai sobre aquesta columna.
# Piso 14C: columnes en l'idioma del llibre de la gestoria -- "Total"
# (brut) es renombra "Base + IVA (€)" i es mou ABANS de %Ret/Retenció
# (es llegeix d'esquerra a dreta com el llibre: brut, retenció, net);
# "Líquid (€)" es renombra "TOTAL FRA. (€)" (el seu llibre defineix
# TOTAL FRA. = base+IVA-retenció) i es queda últim abans d'Estat.
COLUMNAS_DETALLE = [
    "Data", "Núm. factura", "Proveïdor", "NIF", "Base", "%IVA", "Quota",
    "Base + IVA (€)", "% Ret.", "Retenció (€)", "TOTAL FRA. (€)", "Estat",
]
COLUMNA_BASE_DETALLE = "E"
COLUMNA_QUOTA_DETALLE = "G"
COLUMNA_TOTAL_DETALLE = "H"
COLUMNA_RETENCIO_PCT_DETALLE = "I"
COLUMNA_RETENCIO_DETALLE = "J"
COLUMNA_LIQUID_DETALLE = "K"


def celda_fecha(ws, fila, columna, fecha_iso):
    """Piso 13O: objecte datetime real (no text) -- Excel el tracta
    com a DATA de veritat (ordenable, filtrable, llest per a la
    importació a Geyce V2), amb format dd/mm/yyyy. Si fecha_iso es
    buit o no es un ISO valid ("YYYY-MM-DD" -- una OCR erronia, per
    exemple), s'escriu el valor cru tal qual com a text -- mai un
    crash (regla 4: nada falla en silencio, pero tampoc tira avall el
    lot per una data mal llegida)."""
    celda = ws.cell(row=fila, column=columna)
    if fecha_iso:
        try:
            celda.value = datetime.strptime(fecha_iso, "%Y-%m-%d")
            celda.number_format = "dd/mm/yyyy"
            return celda
        except ValueError:
            pass
    celda.value = fecha_iso
    return celda


def _escribir_fila_detalle(ws, fila, nombre, datos, linea, carpeta_original, carpeta_cliente, decision, nombres_ya_mostrados, relleno_forzado=None):
    """Escribe una fila de DETALL (una linia d'IVA). Si relleno_forzado
    es None, el color/estat se deriva de decision/estado (bloque
    FACTURES QUE SUMEN); si no, se usa tal cual (bloque PENDENTS,
    siempre taronja). nombres_ya_mostrados es un set compartido entre
    llamadas de la MISMA escribir_detalle -- Retenció, Total i TOTES les
    altres columnes de nivell factura (Data/Núm/Contrapart/NIF/Estat) son
    por factura, no por linia d'IVA, asi que solo se escriben en la
    primera linia escrita de cada factura: escribir_detalle (Piso 14) les
    fusiona verticalment despres amb un merge_cells -- escribir-les a
    cada fila fundiria un valor per sobre d'un altre en el mateix merge.
    Base/%IVA/Quota (columnes E/F/G) son per linia, sempre."""
    estado = datos.get("estado")
    ruta_original = encontrar_original(carpeta_original, nombre)

    es_abonament = (datos.get("total") or 0) < 0
    tiene_avisos = (
        validar_nif(datos.get("nif_proveedor")) is False
        or validar_nif(datos.get("nif_receptor")) is False
        or not verificar_retencion(datos)
    )

    if relleno_forzado is not None:
        estado_mostrado = estado
        relleno = relleno_forzado
    elif estat_efectiu(decision) == "descartar":
        estado_mostrado = f"DESCARTAT ({decision.get('nota', '')})"
        relleno = RELLENO_DESCARTAT
    elif estat_efectiu(decision) == "aprovar":
        estado_mostrado = f"OK (aprovat manualment per {decision.get('qui', '')}, {decision.get('data', '')})"
        relleno = RELLENO_OK
    else:
        estado_mostrado = estado
        relleno = RELLENO_AVISO if estado == "REVISAR" else RELLENO_OK

    if es_abonament:
        estado_mostrado += " (ABONAMENT)"
    if tiene_avisos:
        estado_mostrado += " [AVÍS]"

    # Piso 11B: distintiu "corregit" -- mismo patron aditivo que
    # ABONAMENT/[AVÍS]. camps_corregits lo escribe validar.py (cirugia
    # minima alli); aqui solo se muestra, ninguna suma cambia.
    camps_corregits = datos.get("camps_corregits") or []
    if camps_corregits:
        detall = "; ".join(f"{c['camp']}: {c['antic']} -> {c['nou']}" for c in camps_corregits)
        estado_mostrado += f" | CORREGIT ({detall})"

    # Piso 13M: mateix patro additiu -- nomes si hi ha hagut alguna
    # decisio revertida (resumen_historial_decisiones ja torna None
    # si no, sense soroll per a la resta).
    resumen_historial = resumen_historial_decisiones(historial_decisiones(carpeta_cliente, nombre))
    if resumen_historial:
        estado_mostrado += f" | HISTORIAL: {resumen_historial}"

    # Piso 13R: mateix patró additiu -- marca visible al DETALL quan la
    # fitxa és exenta=true (el cub propi ja les agrupa, aquesta és la
    # marca a nivell de fila individual).
    if datos.get("exenta"):
        estado_mostrado += " | EXEMPTA"

    # Piso 13X: mateix patró additiu -- fitxa nascuda a "Entrada manual"
    # (app.py), mai passada per extraer_todas.py.
    if datos.get("origen") == "manual":
        estado_mostrado += " | ENTRADA MANUAL"

    # Piso 13Y: mateix patró additiu -- validar.py ha normalitzat el
    # total (el paper duia el líquid, no el brut). El total ja es el
    # brut aquí; la nota amb les dues xifres viu a observaciones.
    if datos.get("total_normalitzat"):
        estado_mostrado += " | TOTAL NORMALITZAT"

    celda_base = ws.cell(row=fila, column=5, value=linea.get("base"))
    celda_base.number_format = FORMATO_MONEDA
    celda_tipo = ws.cell(row=fila, column=6, value=linea.get("tipo_iva"))
    celda_tipo.number_format = FORMATO_PORCENTAJE
    celda_cuota = ws.cell(row=fila, column=7, value=linea.get("cuota"))
    celda_cuota.number_format = FORMATO_MONEDA
    # Piso 14C: ordre nou -- Base+IVA (brut) ABANS de %Ret/Retenció,
    # llegit d'esquerra a dreta com el llibre (brut, retenció, net).
    celda_total = ws.cell(row=fila, column=8)
    celda_total.number_format = FORMATO_MONEDA
    celda_retencio_pct = ws.cell(row=fila, column=9)
    celda_retencio_pct.number_format = FORMATO_PORCENTAJE
    celda_retencio = ws.cell(row=fila, column=10)
    celda_retencio.number_format = FORMATO_MONEDA
    celda_liquid = ws.cell(row=fila, column=11)
    celda_liquid.number_format = FORMATO_MONEDA
    if nombre not in nombres_ya_mostrados:
        celda_fecha(ws, fila, 1, datos.get("fecha_factura"))
        celda_num = ws.cell(row=fila, column=2, value=datos.get("num_factura"))
        if ruta_original:
            asignar_hipervinculo(celda_num, ruta_original)
        # Piso 13K: contrapart (qui NO es el client, per NIF -- validar.py),
        # no "proveedor" a seques -- a rebudes dona el mateix d'abans (el
        # client hi es sempre receptor), a ingressos es la peça que fallava.
        ws.cell(row=fila, column=3, value=datos.get("contrapart_nom"))
        ws.cell(row=fila, column=4, value=datos.get("contrapart_nif"))
        total = datos.get("total") or 0
        celda_total.value = datos.get("total")
        celda_retencio_pct.value = datos.get("retencion_pct") or None
        celda_retencio.value = datos.get("retencion_cuota") or None
        # Piso 13Y/14C: TOTAL FRA. = total - retencio (mai buit -- sense
        # retencio, TOTAL FRA. = Base + IVA). Nomes informatiu, veure
        # nota a COLUMNAS_DETALLE -- cap suma d'aquest fitxer el fa servir.
        celda_liquid.value = total - (datos.get("retencion_cuota") or 0)
        ws.cell(row=fila, column=12, value=estado_mostrado)
        nombres_ya_mostrados.add(nombre)

    for col in range(1, len(COLUMNAS_DETALLE) + 1):
        ws.cell(row=fila, column=col).fill = relleno


# Piso 14: columnes de nivell factura (Data/Núm/Contrapart/NIF/Base+IVA/
# %Ret/Retenció/TOTAL FRA./Estat) -- es fusionen verticalment quan una
# factura te mes d'una línia d'IVA. Base/%IVA/Quota (5, 6, 7) en queden
# fora a proposit: son per línia, mai es fusionen.
COLUMNAS_NIVELL_FACTURA = (1, 2, 3, 4, 8, 9, 10, 11, 12)

BORDE_GRUP_LATERAL = Side(style="thin")


def fusionar_grup_factura(ws, fila_ini, fila_fin):
    """Piso 14: fusiona verticalment les columnes de nivell factura d'un
    grup de files que pertanyen a la MATEIXA factura -- es llegeix UNA
    factura amb N línies i UN total, en lloc de N files que semblen N
    factures. Sense efecte si fila_fin == fila_ini (una sola línia):
    fusionar una cel·la amb ella mateixa no fa res."""
    if fila_fin <= fila_ini:
        return
    for columna in COLUMNAS_NIVELL_FACTURA:
        ws.merge_cells(start_row=fila_ini, end_row=fila_fin, start_column=columna, end_column=columna)


def aplicar_vora_grup(ws, fila_ini, fila_fin, num_columnas):
    """Piso 14: vora de caixa al voltant del grup sencer d'una factura
    (totes les columnes, files fila_ini..fila_fin) -- delimita
    visualment on comença i acaba cada factura, tingui una sola línia
    d'IVA o vàries. Openpyxl no te un "border de rang": cada cel·la del
    perímetre necessita el seu propi Border amb nomes els costats que li
    toquen (les files interiors no en porten cap, perque el grup es
    llegeixi com una sola caixa continua)."""
    for f in range(fila_ini, fila_fin + 1):
        for col in range(1, num_columnas + 1):
            ws.cell(row=f, column=col).border = Border(
                top=BORDE_GRUP_LATERAL if f == fila_ini else None,
                bottom=BORDE_GRUP_LATERAL if f == fila_fin else None,
                left=BORDE_GRUP_LATERAL if col == 1 else None,
                right=BORDE_GRUP_LATERAL if col == num_columnas else None,
            )


def escribir_detalle(ws, fila, titulo, facturas, carpeta_original, carpeta_cliente, decisiones):
    """Piso 12A: DETALL en dos bloques -- "FACTURES QUE SUMEN" (las
    mismas que hoy cuentan en sumar_bloque: OK crudo o decision
    aprovar, nunca descartades); y debajo "PENDENTS -- NO SUMEN" (las
    REVISAR sin decision, en taronja, con el mismo detalle de linias,
    para auditar sin que sumen). Las DESCARTADES ya tienen su propio
    bloque (DESCARTATS, con nota/qui/data) y no se repiten aqui.

    Piso 14: ordenado por FACTURA (data, 13O) -- las N linias de una
    misma factura quedan SIEMPRE contiguas (fusionar_grup_factura las
    fusiona verticalmente despues), aunque tengan tipos de IVA
    distintos (el "cas rei": 10%+21% en la misma factura). Antes se
    ordenaba por tipo primero para que cada cubo fuera un rango
    contiguo -- ya no es asi: rango_tipos ahora es una LISTA de filas
    exactas por tipo (pueden estar dispersas entre facturas),
    escribir_formulas_bloque/escribir_formulas_resultat escriben
    =SUM(fila1,fila2,...) en vez de =SUM(ini:fin).

    Devuelve (fila, rangos) -- rangos = {"bloque": (ini, fin) o None,
    "tipos": {tipo: [fila, fila, ...]}} con las filas EXACTAS de
    "FACTURES QUE SUMEN", que escribir_formulas_bloque/
    escribir_formulas_resultat necesitan para poder escribir sus
    =SUM()."""
    facturas_suman = []
    facturas_pendientes = []
    for nombre, datos in facturas:
        decision = decisiones.get(nombre)
        if estat_efectiu(decision) == "descartar":
            continue
        cuenta = datos.get("estado") == "OK" or estat_efectiu(decision) == "aprovar"
        if cuenta:
            facturas_suman.append((nombre, datos))
        else:
            facturas_pendientes.append((nombre, datos))

    # Piso 13K: la columna 3 mostra la contrapart (validar.py) -- a
    # INGRESSOS es diu "Client" (la contrapart hi es qui compra), a
    # DESPESES es queda "Proveïdor" (la contrapart hi es qui ven).
    columnas = list(COLUMNAS_DETALLE)
    if titulo == "DETALL INGRESSOS":
        columnas[2] = "Client"

    for col in range(1, len(COLUMNAS_DETALLE) + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value=titulo)
    fila += 1

    ws.cell(row=fila, column=1, value="FACTURES QUE SUMEN").font = ESTILO_ENCABEZADO
    fila += 1
    for col, texto in enumerate(columnas, start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    # Piso 13R/13O: ordre per data -- les fitxes sense fecha_factura
    # cauen sempre al final. Piso 14: ja NO s'ordena per tipus primer --
    # cada factura escriu les seves N línies seguides, tingui un o
    # varis tipus d'IVA.
    facturas_suman.sort(key=lambda nd: nd[1].get("fecha_factura") or "9999-99-99")

    nombres_ya_mostrados = set()
    rango_tipos = {}
    fila_inicio_bloque = fila if facturas_suman else None
    for nombre, datos in facturas_suman:
        decision = decisiones.get(nombre)
        fila_inicio_factura = fila
        for linea in (datos.get("lineas_iva") or [{}]):
            tipo = linea.get("tipo_iva")
            # Piso 14B: exenta es un camp DE LÍNIA -- cada línia va al
            # seu cub, encara que la factura sigui mixta.
            clave_tipo = "exenta" if linea.get("exenta") else (tipo if tipo in TIPOS_IVA else "otros")
            rango_tipos.setdefault(clave_tipo, []).append(fila)
            _escribir_fila_detalle(ws, fila, nombre, datos, linea, carpeta_original, carpeta_cliente, decision, nombres_ya_mostrados)
            fila += 1
        fila_fin_factura = fila - 1
        fusionar_grup_factura(ws, fila_inicio_factura, fila_fin_factura)
        aplicar_vora_grup(ws, fila_inicio_factura, fila_fin_factura, len(COLUMNAS_DETALLE))
    fila_fin_bloque = (fila - 1) if facturas_suman else None

    rangos = {"bloque": (fila_inicio_bloque, fila_fin_bloque) if facturas_suman else None, "tipos": rango_tipos}
    fila += 1  # fila en blanco

    if facturas_pendientes:
        # Piso 13R: ordre per data -- sense rangs ni formules que en
        # depenguin (aquest bloc no suma res), un sort pla n'hi ha prou.
        facturas_pendientes.sort(key=lambda np: np[1].get("fecha_factura") or "9999-99-99")
        ws.cell(row=fila, column=1, value="PENDENTS — NO SUMEN").font = ESTILO_ENCABEZADO
        for col in range(1, len(COLUMNAS_DETALLE) + 1):
            ws.cell(row=fila, column=col).fill = RELLENO_AVISO
        fila += 1
        for col, texto in enumerate(columnas, start=1):
            ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
        fila += 1
        nombres_ya_mostrados_pendents = set()
        for nombre, datos in facturas_pendientes:
            fila_inicio_factura = fila
            for linea in (datos.get("lineas_iva") or [{}]):
                _escribir_fila_detalle(
                    ws, fila, nombre, datos, linea, carpeta_original, carpeta_cliente, None,
                    nombres_ya_mostrados_pendents, relleno_forzado=RELLENO_AVISO,
                )
                fila += 1
            fila_fin_factura = fila - 1
            fusionar_grup_factura(ws, fila_inicio_factura, fila_fin_factura)
            aplicar_vora_grup(ws, fila_inicio_factura, fila_fin_factura, len(COLUMNAS_DETALLE))

    return fila + 1, rangos


def cargar_manifiestos(carpeta_lotes_procesados):
    """Lee todos los *_manifiesto.json de lotes_procesados/. Devuelve
    lista de (nombre_lote_pdf, documentos). Los lotes procesados antes
    del piso 6A no tienen manifiesto -- quedan fuera, asumido."""
    manifiestos = []
    if not os.path.isdir(carpeta_lotes_procesados):
        return manifiestos
    for nombre in sorted(os.listdir(carpeta_lotes_procesados)):
        if not nombre.endswith("_manifiesto.json"):
            continue
        nombre_lote_pdf = nombre[: -len("_manifiesto.json")] + ".pdf"
        with open(os.path.join(carpeta_lotes_procesados, nombre), encoding="utf-8") as f:
            documentos = json.load(f)
        manifiestos.append((nombre_lote_pdf, documentos))
    return manifiestos


# Piso 13K: els dos molls possibles d'un lot ja processat -- cal provar
# tots dos perque un manifest de ruido no diu de quin moll venia.
CARPETAS_LOTES_PROCESADOS = ["rebudes/lotes_procesados", "apartados/lotes_vendes_procesados"]


def _ruta_lote_procesat(carpeta_cliente, nombre_lote):
    """Piso 13K: on sigui que hagi acabat el PDF sencer ja processat
    (moll de compres o de vendes) -- None si no es troba a cap dels
    dos (no hauria de passar, pero l'enllaç simplement no es crea)."""
    for subcarpeta in CARPETAS_LOTES_PROCESADOS:
        ruta = os.path.join(carpeta_cliente, subcarpeta, nombre_lote)
        if os.path.exists(ruta):
            return ruta
    return None


def avisos_consistencia(facturas):
    """Agrupa por nombre de proveedor (no por nif_proveedor: un NIF mal
    leido por OCR varia de una factura a otra del mismo proveedor real,
    mientras que el nombre suele salir igual -- ver el caso del Celler).
    Devuelve una entrada por proveedor que use mas de un tipo de IVA
    distinto entre sus facturas -- OK y REVISAR juntas, es solo
    informativo, no cambia ningun estado."""
    por_proveedor = {}
    for nombre, datos in facturas:
        proveedor = datos.get("proveedor")
        if proveedor is None:
            continue
        entrada = por_proveedor.setdefault(proveedor, {"nifs": set(), "tipos": set(), "archivos": []})
        nif = datos.get("nif_proveedor")
        if nif is not None:
            entrada["nifs"].add(nif)
        for linea in datos.get("lineas_iva") or []:
            tipo = linea.get("tipo_iva")
            if tipo is not None:
                entrada["tipos"].add(tipo)
        entrada["archivos"].append(nombre)

    avisos = []
    for proveedor, info in sorted(por_proveedor.items()):
        if len(info["tipos"]) > 1:
            avisos.append((proveedor, sorted(info["nifs"]), sorted(info["tipos"]), info["archivos"]))
    return avisos


def avisos_verificacion(facturas):
    """Devuelve lista de (nombre, datos, motivo) para fichas cuyo NIF
    no supera la letra de control o cuya retencion no cuadra. No
    cambia ningun estado ni suma -- son avisos aparte."""
    avisos = []
    for nombre, datos in facturas:
        for campo, etiqueta in [("nif_proveedor", "proveïdor"), ("nif_receptor", "receptor")]:
            valor = datos.get(campo)
            if validar_nif(valor) is False:
                avisos.append((nombre, datos, f"NIF no supera la validació de la lletra ({etiqueta}: {valor})"))
        if not verificar_retencion(datos):
            avisos.append((nombre, datos, "la retenció calculada no quadra amb la retenció indicada"))
    return avisos


def escribir_avisos(ws, fila, carpeta_cliente, gastos, ingresos, origen_gastos, origen_ingressos):
    # a) DOCUMENTS APARTATS -- albarans que trocear.py aparta i ningu mes processa
    for col in range(1, 3):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value="DOCUMENTS APARTATS")
    fila += 1
    for col, texto in enumerate(["Fitxer", "Justificació"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    carpeta_albarans = f"{carpeta_cliente}/apartados/albarans"
    nombres_albaran = []
    if os.path.isdir(carpeta_albarans):
        nombres_albaran = sorted(
            n for n in os.listdir(carpeta_albarans) if n.lower().endswith(EXTENSIONES_ORIGINAL)
        )

    if not nombres_albaran:
        ws.cell(row=fila, column=1, value="sense documents apartats")
        fila += 1
    else:
        for nombre in nombres_albaran:
            celda = ws.cell(row=fila, column=1, value=nombre)
            asignar_hipervinculo(celda, os.path.join(carpeta_albarans, nombre))
            celda_justificacion = ws.cell(
                row=fila,
                column=2,
                value="Albarà: no es comptabilitza — la factura posterior n'agrupa els albarans. Verificar que aquesta factura ha arribat",
            )
            celda_justificacion.alignment = AJUSTE_TEXTO
            fila += 1
    fila += 1

    # b) PAGINES DESCARTADES COM A SOROLL -- llegides dels manifests de trocear.py
    for col in range(1, 4):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value="PÀGINES DESCARTADES COM A SOROLL")
    fila += 1
    for col, texto in enumerate(["Lot", "Pàgines", "Pista"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    # Piso 13K: dos molls (compres i vendes) tenen cada un el seu propi
    # lotes_procesados/ -- es llegeixen tots dos perque cap pagina de
    # soroll quedi invisible nomes per venir del moll de vendes.
    manifiestos = (
        cargar_manifiestos(f"{carpeta_cliente}/rebudes/lotes_procesados")
        + cargar_manifiestos(f"{carpeta_cliente}/apartados/lotes_vendes_procesados")
    )
    filas_ruido = [
        (nombre_lote, doc)
        for nombre_lote, documentos in manifiestos
        for doc in documentos
        if doc.get("tipo") == "ruido"
    ]

    if not manifiestos:
        ws.cell(row=fila, column=1, value="sense dades de lots anteriors")
        fila += 1
    elif not filas_ruido:
        ws.cell(row=fila, column=1, value="sense pàgines de soroll")
        fila += 1
    else:
        for nombre_lote, doc in filas_ruido:
            celda = ws.cell(row=fila, column=1, value=nombre_lote)
            ruta_lote = _ruta_lote_procesat(carpeta_cliente, nombre_lote)
            if ruta_lote:
                asignar_hipervinculo(celda, ruta_lote)
            ws.cell(row=fila, column=2, value=f"p{doc['pagina_inicio']}-{doc['pagina_fin']}")
            ws.cell(row=fila, column=3, value=doc.get("emisor_pista"))
            fila += 1
    fila += 1

    # c) AVISOS DE CONSISTENCIA -- mateix proveidor, mes d'un tipus d'IVA
    for col in range(1, 6):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value="AVISOS DE CONSISTÈNCIA")
    fila += 1
    for col, texto in enumerate(["Flux", "Proveïdor", "NIF", "Tipus d'IVA", "Fitxers afectats"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    filas_consistencia = [("DESPESA", a) for a in avisos_consistencia(gastos)] + [
        ("INGRÉS", a) for a in avisos_consistencia(ingresos)
    ]

    if not filas_consistencia:
        ws.cell(row=fila, column=1, value="sense avisos de consistència")
        fila += 1
    else:
        for flujo, (proveedor, nifs, tipos, archivos) in filas_consistencia:
            ws.cell(row=fila, column=1, value=flujo)
            ws.cell(row=fila, column=2, value=proveedor)
            ws.cell(row=fila, column=3, value=" / ".join(nifs))
            ws.cell(row=fila, column=4, value="{" + ", ".join(str(t) for t in tipos) + "}")
            celda_archivos = ws.cell(row=fila, column=5, value="; ".join(archivos))
            celda_archivos.alignment = AJUSTE_TEXTO
            fila += 1

    # d) ERRORS -- archivos presentes sin ficha extraida (no tienen fecha, no van en ningun trimestre)
    for col in range(1, 3):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_ERROR
    ws.cell(row=fila, column=1, value="ERRORS")
    fila += 1
    for col, texto in enumerate(["Fitxer", "Motiu"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    errores_gastos = detectar_errores(listar_archivos_rebudes(origen_gastos), f"{carpeta_cliente}/rebudes/extraidas")
    errores_ingresos = detectar_errores(
        listar_archivos_planos(origen_ingressos), f"{carpeta_cliente}/apartados/ingressos_extraidas"
    )
    filas_error = [("DESPESA", r) for r in errores_gastos] + [("INGRÉS", r) for r in errores_ingresos]

    if not filas_error:
        ws.cell(row=fila, column=1, value="sense errors")
        fila += 1
    else:
        for flujo, ruta in filas_error:
            celda = ws.cell(row=fila, column=1, value=f"[{flujo}] {os.path.basename(ruta)}")
            asignar_hipervinculo(celda, ruta)
            celda_motivo = ws.cell(row=fila, column=2, value=motivo_error(ruta))
            celda_motivo.alignment = AJUSTE_TEXTO
            for col in range(1, 3):
                ws.cell(row=fila, column=col).fill = RELLENO_ERROR
            fila += 1
    fila += 1

    # e) AVISOS DE VERIFICACIO -- letra de NIF y cuadre de retencion, no cambian estado
    for col in range(1, 3):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_BLOQUE
    ws.cell(row=fila, column=1, value="AVISOS DE VERIFICACIÓ")
    fila += 1
    for col, texto in enumerate(["Fitxer", "Motiu"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    filas_verificacion = (
        [(origen_gastos, nombre, datos, motivo) for nombre, datos, motivo in avisos_verificacion(gastos)]
        + [(origen_ingressos, nombre, datos, motivo) for nombre, datos, motivo in avisos_verificacion(ingresos)]
    )
    if not filas_verificacion:
        ws.cell(row=fila, column=1, value="sense avisos de verificació")
        fila += 1
    else:
        for carpeta_origen, nombre, datos, motivo in filas_verificacion:
            celda = ws.cell(row=fila, column=1, value=nombre)
            ruta_original = encontrar_original(carpeta_origen, nombre)
            if ruta_original:
                asignar_hipervinculo(celda, ruta_original)
            celda_motivo = ws.cell(row=fila, column=2, value=motivo)
            celda_motivo.alignment = AJUSTE_TEXTO
            fila += 1

    return (
        len(nombres_albaran), len(filas_ruido), len(filas_consistencia), bool(manifiestos),
        len(filas_error), len(filas_verificacion),
    )


def escribir_pendientes(ws, fila, pendientes, carpeta_cliente):
    for col in range(1, 4):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_AVISO
    ws.cell(row=fila, column=1, value="PENDENT DE REVISIÓ")
    fila += 1
    for col, texto in enumerate(["Fitxer", "Tipus", "Motius"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    for nombre, datos, tipo_bloque, carpeta_original in pendientes:
        celda = ws.cell(row=fila, column=1, value=nombre)
        ruta_original = encontrar_original(carpeta_original, nombre)
        if ruta_original:
            asignar_hipervinculo(celda, ruta_original)
        else:
            print(f"AVISO: no se encontró el original de {nombre}")
        ws.cell(row=fila, column=2, value=tipo_bloque)
        motivos_traducidos = [traducir_motivo(m) for m in (datos.get("motivos") or [])]
        celda_motivos = ws.cell(row=fila, column=3, value="; ".join(motivos_traducidos))
        celda_motivos.alignment = AJUSTE_TEXTO
        fila += 1

    return fila + 1


def escribir_descartados(ws, fila, descartados, carpeta_cliente):
    """Facturas con decisions.csv accion=descartar -- nunca cuentan en
    ninguna suma, se listan aparte con su nota."""
    for col in range(1, 4):
        celda = ws.cell(row=fila, column=col)
        celda.font = ESTILO_ENCABEZADO
        celda.fill = RELLENO_DESCARTAT
    ws.cell(row=fila, column=1, value="DESCARTATS")
    fila += 1
    for col, texto in enumerate(["Fitxer", "Tipus", "Nota"], start=1):
        ws.cell(row=fila, column=col, value=texto).font = ESTILO_ENCABEZADO
    fila += 1

    for nombre, datos, tipo_bloque, carpeta_original, decision in descartados:
        celda = ws.cell(row=fila, column=1, value=nombre)
        ruta_original = encontrar_original(carpeta_original, nombre)
        if ruta_original:
            asignar_hipervinculo(celda, ruta_original)
        ws.cell(row=fila, column=2, value=tipo_bloque)
        nota = decision.get("nota") or ""
        qui = decision.get("qui") or ""
        data = decision.get("data") or ""
        celda_nota = ws.cell(row=fila, column=3, value=f"{nota} ({qui}, {data})")
        celda_nota.alignment = AJUSTE_TEXTO
        fila += 1

    return fila + 1


for fila_cliente in leer_clientes():
    carpeta = fila_cliente["carpeta"]
    carpeta_cliente = f"clientes/{carpeta}"

    gastos = cargar_validadas(f"{carpeta_cliente}/rebudes/validadas")
    ingresos = cargar_validadas(f"{carpeta_cliente}/apartados/ingressos_validadas")

    if not gastos and not ingresos:
        continue

    origen_gastos = f"{carpeta_cliente}/rebudes"
    origen_ingressos = f"{carpeta_cliente}/{RUTAS_ORIGEN_INGRESSOS_PERSONALIZADAS.get(carpeta, 'apartados/ingressos')}"

    decisiones = cargar_decisiones(carpeta_cliente)
    nombres_validos = {n for n, _ in gastos} | {n for n, _ in ingresos}
    for archivo in decisiones:
        if archivo not in nombres_validos:
            print(f"AVISO: decisions.csv de {carpeta} referencia un archivo que no existe entre las validadas: {archivo}")

    # Agrupar por trimestre (o SIN FECHA si una REVISAR no trae fecha_factura)
    trimestres = {}
    for nombre, datos in gastos:
        t = trimestre_de(datos.get("fecha_factura")) or "SIN FECHA"
        trimestres.setdefault(t, {"gastos": [], "ingresos": []})["gastos"].append((nombre, datos))
    for nombre, datos in ingresos:
        t = trimestre_de(datos.get("fecha_factura")) or "SIN FECHA"
        trimestres.setdefault(t, {"gastos": [], "ingresos": []})["ingresos"].append((nombre, datos))

    wb = Workbook()
    wb.remove(wb.active)

    orden = sorted(t for t in trimestres if t != "SIN FECHA") + (
        ["SIN FECHA"] if "SIN FECHA" in trimestres else []
    )

    comprobaciones_por_hoja = {}

    for trimestre in orden:
        datos_trimestre = trimestres[trimestre]
        nombre_hoja = "SENSE DATA" if trimestre == "SIN FECHA" else trimestre
        ws = wb.create_sheet(nombre_hoja)
        configurar_impresion(ws)
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 14
        ws.column_dimensions["L"].width = 12

        fila = escribir_titulo(ws, fila_cliente["nombre"], fila_cliente["nif"])
        pendientes = []
        descartados_totales = []

        if trimestre != "SIN FECHA":
            # Piso 12A: 1a pasada -- estructura de RESULTAT/DESPESES/INGRESSOS,
            # celdas numericas en blanco (se rellenan en la 3a pasada, una vez
            # se conoce el rango exacto de filas de DETALL).
            sumas_g, total_g, retencion_g, revisar_g, descartados_g = sumar_bloque(datos_trimestre["gastos"], decisiones)
            sumas_i, total_i, retencion_i, revisar_i, descartados_i = sumar_bloque(datos_trimestre["ingresos"], decisiones)

            fila, info_resultat = escribir_resultat_estructura(ws, fila, sumas_g, sumas_i)
            # Piso 13S: DESPESES guanya la mateixa fila de retencions que
            # INGRESSOS ja tenia -- les del lloguer (model 115), etiqueta
            # propia perque no es confongui amb la d'INGRESSOS.
            fila, info_g = escribir_bloque(
                ws, fila, "DESPESES", sumas_g, con_retencion=True, etiqueta_retencion="Σ RETENCIONS SUPORTADES",
            )
            fila, info_i = escribir_bloque(ws, fila, "INGRESSOS", sumas_i, con_retencion=True)

            n_ok_g = len(datos_trimestre["gastos"]) - len(revisar_g) - len(descartados_g)
            n_ok_i = len(datos_trimestre["ingresos"]) - len(revisar_i) - len(descartados_i)
            print(
                f"{carpeta} / {trimestre}: GASTOS total={total_g:.2f} ({n_ok_g} OK, {len(revisar_g)} a revisar) | "
                f"INGRESOS total={total_i:.2f} ({n_ok_i} OK, {len(revisar_i)} a revisar)"
            )
        else:
            # Sin fecha no hay nada que sumar por trimestre -- solo se lista para revisar
            _, _, _, revisar_g, descartados_g = sumar_bloque(datos_trimestre["gastos"], decisiones)
            _, _, _, revisar_i, descartados_i = sumar_bloque(datos_trimestre["ingresos"], decisiones)
            print(f"{carpeta} / SIN FECHA: {len(revisar_g) + len(revisar_i)} a revisar (sin fecha_factura)")

        for nombre, datos in revisar_g:
            pendientes.append((nombre, datos, "DESPESA", f"{carpeta_cliente}/rebudes"))
        for nombre, datos in revisar_i:
            # Piso 13J: origen_ingressos YA porta carpeta_cliente (linea de
            # mas abajo donde se define) -- prependerlo otra vez rompia
            # encontrar_original para cualquier ingres PENDENT de un client
            # amb origen personalitzat (nomes davinstal fins ara). Bug latent
            # descobert en verificar migrar_lot.py -- mai havia passat en
            # producció perque tots els ingressos de davinstal eren OK.
            pendientes.append((nombre, datos, "INGRÉS", origen_ingressos))

        # 2a pasada: PENDENT DE REVISIÓ y DESCARTATS, igual que siempre.
        if pendientes:
            # Piso 13R: ordre per data -- aquesta taula no suma res, cap
            # rang ni formula en depèn.
            pendientes.sort(key=lambda p: p[1].get("fecha_factura") or "9999-99-99")
            fila = escribir_pendientes(ws, fila, pendientes, carpeta_cliente)

        for nombre, datos, decision in descartados_g:
            descartados_totales.append((nombre, datos, "DESPESA", f"{carpeta_cliente}/rebudes", decision))
        for nombre, datos, decision in descartados_i:
            descartados_totales.append((nombre, datos, "INGRÉS", origen_ingressos, decision))

        if descartados_totales:
            fila = escribir_descartados(ws, fila, descartados_totales, carpeta_cliente)

        if trimestre != "SIN FECHA":
            # 3a pasada: DETALL (anota los rangos de fila de "FACTURES QUE
            # SUMEN") y, con esos rangos ya conocidos, se rellenan las celdas
            # que la 1a pasada dejo en blanco con formulas =SUM().
            fila, rango_g = escribir_detalle(
                ws, fila, "DETALL DESPESES", datos_trimestre["gastos"],
                f"{carpeta_cliente}/rebudes", carpeta_cliente, decisiones,
            )
            fila, rango_i = escribir_detalle(
                ws, fila, "DETALL INGRESSOS", datos_trimestre["ingresos"],
                origen_ingressos, carpeta_cliente, decisiones,
            )

            comprobaciones = []
            comprobaciones += escribir_formulas_bloque(ws, info_g, rango_g, sumas_g, total_g, retencion_g)
            comprobaciones += escribir_formulas_bloque(ws, info_i, rango_i, sumas_i, total_i, retencion_i)
            comprobaciones += escribir_formulas_resultat(ws, info_resultat, rango_g, rango_i, sumas_g, sumas_i)
            comprobaciones_por_hoja[nombre_hoja] = comprobaciones

    ws_avisos = wb.create_sheet("AVISOS")
    configurar_impresion(ws_avisos)
    ws_avisos.column_dimensions["A"].width = 55
    ws_avisos.column_dimensions["B"].width = 70
    ws_avisos.column_dimensions["C"].width = 20
    ws_avisos.column_dimensions["D"].width = 20
    ws_avisos.column_dimensions["E"].width = 70

    fila = escribir_titulo(ws_avisos, fila_cliente["nombre"], fila_cliente["nif"])
    n_albaranes, n_ruido, n_consistencia, hay_manifiestos, n_errores, n_verificacion = escribir_avisos(
        ws_avisos, fila, carpeta_cliente, gastos, ingresos, origen_gastos, origen_ingressos
    )
    detalle_ruido = f"{n_ruido} páginas de ruido" if hay_manifiestos else "sin datos de lotes anteriores"
    print(
        f"{carpeta} / AVISOS: {n_albaranes} documentos apartados, {detalle_ruido}, "
        f"{n_consistencia} avisos de consistencia, {n_errores} errores, {n_verificacion} avisos de verificación"
    )

    ruta_excel = f"{carpeta_cliente}/sumatorios_2026.xlsx"
    # Piso 13G/13W: a Windows, un Excel obert bloqueja el fitxer --
    # es captura per client (regla 4: el lot mai mor) en comptes de
    # tirar avall tot el batch, i mai deixa un fitxer a mig escriure
    # (un altre cami cap a "reparar"). Piso 13W: abans de rendir-se,
    # reintenta amb paciencia (guardar_amb_reintents) -- un PermissionError
    # no vol dir necessariament que Excel el tingui obert (vegeu la
    # funcio), aixi que el missatge final mai ho afirma com a cert.
    ok, error = guardar_amb_reintents(lambda: wb.save(ruta_excel))
    if not ok:
        print(
            f"AVISO: {carpeta} / Excel: ✗ bloquejat -- no s'ha pogut escriure {ruta_excel} "
            f"despres de 6 intents (~10s): {type(error).__name__}: {error}. Tanca l'Excel "
            f"d'aquest client si el tens obert, o torna-ho a provar en uns segons."
        )
        continue
    print(f"{carpeta} / Excel: ✓ actualitzat ({ruta_excel})")
    if comprobaciones_por_hoja:
        # Piso 13J: mismo motivo que el PermissionError de wb.save() --
        # reescribe el archivo entero, y en Windows podria estar
        # bloqueado si alguien lo abrio justo entre el guardado y aqui.
        ok_inj, error_inj = guardar_amb_reintents(
            lambda: inyectar_valores_cacheados(ruta_excel, comprobaciones_por_hoja)
        )
        if not ok_inj:
            print(
                f"AVISO: {carpeta} / Excel: ✗ valors calculats no injectats -- no s'ha pogut "
                f"reescriure {ruta_excel} despres de 6 intents (~10s): "
                f"{type(error_inj).__name__}: {error_inj} (Excel 2007 podria seguir demanant "
                f"reparar). Tanca l'Excel d'aquest client si el tens obert, o torna-ho a "
                f"provar en uns segons."
            )
    verificar_enlaces_excel(ruta_excel, carpeta_cliente)
    if comprobaciones_por_hoja:
        n_formulas = verificar_formulas(ruta_excel, comprobaciones_por_hoja)
        print(f"{carpeta} / excel: Fórmules verificades: {n_formulas}")
