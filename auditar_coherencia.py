"""Piso 14E: VERIFICADOR DE COHERÈNCIA -- l'auditor que comprova que
cada mutació (moure, retirar, reactivar, corregir, decidir, desfer,
manual) ha deixat el món SENCER. Codi pur, zero API, zero streamlit.

Recorre tots els documents coneguts d'un client (fitxes, originals,
retirats, decisions, moviments, manuals, hashos) i verifica les QUATRE
capes de cadascun:

  FÍSICA  -- l'arxiu és exactament on el seu estat mana: mai a dues
             carpetes excloents alhora; el retirat al cementiri, el
             reactivat fora.
  FITXA   -- tota fitxa activa té el seu original (o origen="manual",
             arxiu nul legítim); cap fitxa extreta sense validar; cada
             moviment de flux/client va moure fitxa I arxiu junts.
  LLIBRES -- hashos.csv resol a arxius reals o a morts explicades
             (retirat/mogut/destruït); decisions.csv apunta a fitxes
             que existeixen.
  VISTES  -- el que estat_efectiu declara viu és al detall de l'Excel;
             el mort, fora. Si l'Excel és més vell que l'última mutació
             dels llibres, es diu UNA vegada ("falta Recalcular") i no
             es sorolla amb el detall (el semàfor ja ho persegueix).

Sortida: divergències amb detall exacte (document, capa, esperat,
trobat). 0 divergències = sistema coherent.

CLI:
  python3 auditar_coherencia.py [carpetes...]     exit 1 si divergeix
  python3 auditar_coherencia.py --informatiu ...  exit 0 sempre; les
      divergències surten amb prefix "AVISO: coherència" perquè el
      detector d'AVISO de Recalcular/Processar s'encengui sol i la
      cadena mai mori per una divergència (regla 4).

Els helpers compartits es dupliquen aquí a consciència, com fan
sumar.py/informe.py/verificar_conservacio.py -- cada màquina es val
per si sola.
"""

import csv
import json
import os
import sys
import unicodedata
from datetime import datetime
from urllib.parse import unquote

from openpyxl import load_workbook


def nfc(nom):
    """macOS guarda els noms de fitxer en NFD (accents descompostos) i
    APFS els tracta com a equivalents als NFC en tota cerca de ruta --
    per això cap màquina del pipeline se n'adona. Però una comparació
    de strings byte a byte (aquest auditor) sí: TOT nom es normalitza
    a NFC abans de comparar, mai una divergència per un accent."""
    return unicodedata.normalize("NFC", nom or "")

RAIZ = os.path.dirname(os.path.abspath(__file__))

EXTENSIONES_ORIGINAL = (".pdf", ".jpg", ".jpeg", ".png")
SUBCARPETAS_RESERVADAS = {"extraidas", "validadas", "procesadas", "lotes_escaneados", "lotes_procesados"}
SUBCARPETAS_NO_ORIGINALES = {"extraidas", "validadas", "lotes_escaneados", "lotes_procesados"}
RUTAS_ORIGEN_INGRESSOS_PERSONALIZADAS = {"davinstal": "Emeses/davinstal"}

SECCIONS_DETALL = {"FACTURES QUE SUMEN": "sumen", "PENDENTS — NO SUMEN": "pendents", "DESCARTATS": "descartats"}
SECCIONS_QUE_TANQUEN = {
    "RESULTAT DEL TRIMESTRE", "DESPESES", "INGRESSOS",
    "DETALL DESPESES", "DETALL INGRESSOS",
    "PENDENT DE REVISIÓ", "DESCARTATS",
}


def leer_clientes():
    ruta = os.path.join(RAIZ, "clientes", "clientes.csv")
    if not os.path.exists(ruta):
        return []
    with open(ruta, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def cargar_validadas(carpeta):
    facturas = []
    if not os.path.isdir(carpeta):
        return facturas
    for nombre in sorted(os.listdir(carpeta)):
        if not nombre.lower().endswith(".json"):
            continue
        with open(os.path.join(carpeta, nombre), encoding="utf-8") as f:
            facturas.append((nfc(nombre), json.load(f)))
    return facturas


def cargar_decisiones(carpeta_cliente):
    ruta = os.path.join(carpeta_cliente, "decisions.csv")
    decisiones = {}
    if not os.path.exists(ruta):
        return decisiones
    with open(ruta, encoding="utf-8") as f:
        for fila in csv.DictReader(f):
            archivo = nfc(fila.get("archivo"))
            if not archivo:
                continue
            if fila.get("accion") == "revertir":
                decisiones.pop(archivo, None)
            else:
                decisiones[archivo] = fila
    return decisiones


def estat_efectiu(decision):
    return decision.get("accion") if decision else None


def encontrar_original(carpeta_origen, nombre_json):
    base = os.path.splitext(nombre_json)[0]
    for ext in EXTENSIONES_ORIGINAL:
        ruta = os.path.join(carpeta_origen, base + ext)
        if os.path.exists(ruta):
            return ruta
    if os.path.isdir(carpeta_origen):
        for nombre_sub in sorted(os.listdir(carpeta_origen)):
            ruta_sub = os.path.join(carpeta_origen, nombre_sub)
            if not os.path.isdir(ruta_sub) or nombre_sub.lower() in SUBCARPETAS_NO_ORIGINALES:
                continue
            for ext in EXTENSIONES_ORIGINAL:
                ruta = os.path.join(ruta_sub, base + ext)
                if os.path.exists(ruta):
                    return ruta
    return None


def trimestre_de(fecha):
    if not fecha:
        return None
    try:
        mes = int(fecha[5:7])
    except (ValueError, IndexError):
        return None
    return f"{(mes - 1) // 3 + 1}T"


def leer_csv(ruta):
    if not os.path.exists(ruta):
        return []
    with open(ruta, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def listar_originals(carpeta):
    """Arxius originals PLANS d'una carpeta (mai subcarpetes)."""
    if not os.path.isdir(carpeta):
        return []
    return [
        nfc(n) for n in sorted(os.listdir(carpeta))
        if n.lower().endswith(EXTENSIONES_ORIGINAL) and os.path.isfile(os.path.join(carpeta, n))
    ]


def listar_originals_rebudes(carpeta_rebudes):
    """entrada/ i subcarpetes germanes no reservades (proveïdors),
    mateix criteri que listar_archivos_rebudes d'app.py."""
    noms = []
    if not os.path.isdir(carpeta_rebudes):
        return noms
    for nombre in sorted(os.listdir(carpeta_rebudes)):
        ruta = os.path.join(carpeta_rebudes, nombre)
        if not os.path.isdir(ruta) or nombre.lower() in SUBCARPETAS_RESERVADAS:
            continue
        noms += [nfc(n) for n in sorted(os.listdir(ruta)) if n.lower().endswith(EXTENSIONES_ORIGINAL)]
    return noms


def divergencia(document, capa, esperat, trobat):
    return {"document": document, "capa": capa, "esperat": esperat, "trobat": trobat}


def data_mutacio_mes_nova(carpeta_cliente):
    """La data més nova de tots els llibres de mutacions -- per saber
    si l'Excel és més vell que l'última mutació (VISTES en fred)."""
    mes_nova = 0
    fonts = [
        os.path.join(carpeta_cliente, "decisions.csv"),
        os.path.join(carpeta_cliente, "moviments_flux.csv"),
        os.path.join(carpeta_cliente, "entrades_manuals.csv"),
        os.path.join(carpeta_cliente, "errors_retirats", "registre.csv"),
        os.path.join(carpeta_cliente, "correccions.csv"),
    ]
    for ruta in fonts:
        for fila in leer_csv(ruta):
            try:
                marca = datetime.strptime(fila.get("data", ""), "%Y-%m-%d %H:%M:%S").timestamp()
            except ValueError:
                continue
            mes_nova = max(mes_nova, marca)
    return mes_nova


def auditar_client(carpeta):
    carpeta_cliente = os.path.join(RAIZ, "clientes", carpeta)
    divergencies = []

    origen_ingressos_rel = RUTAS_ORIGEN_INGRESSOS_PERSONALIZADAS.get(carpeta, "apartados/ingressos")
    carpeta_rebudes = os.path.join(carpeta_cliente, "rebudes")
    carpeta_ingressos = os.path.join(carpeta_cliente, *origen_ingressos_rel.split("/"))
    carpeta_procesadas = os.path.join(carpeta_rebudes, "procesadas")
    carpeta_retirats = os.path.join(carpeta_cliente, "errors_retirats")

    validadas_g = dict(cargar_validadas(os.path.join(carpeta_rebudes, "validadas")))
    validadas_i = dict(cargar_validadas(os.path.join(carpeta_cliente, "apartados", "ingressos_validadas")))
    extraidas_g = {
        nfc(n) for n in os.listdir(os.path.join(carpeta_rebudes, "extraidas"))
        if n.lower().endswith(".json")
    } if os.path.isdir(os.path.join(carpeta_rebudes, "extraidas")) else set()
    extraidas_i = {
        nfc(n) for n in os.listdir(os.path.join(carpeta_cliente, "apartados", "ingressos_extraidas"))
        if n.lower().endswith(".json")
    } if os.path.isdir(os.path.join(carpeta_cliente, "apartados", "ingressos_extraidas")) else set()

    decisiones = cargar_decisiones(carpeta_cliente)
    registre_retirats = leer_csv(os.path.join(carpeta_retirats, "registre.csv"))
    moviments = leer_csv(os.path.join(carpeta_cliente, "moviments_flux.csv"))
    hashos = leer_csv(os.path.join(carpeta_cliente, "hashos.csv"))

    # Ubicacions físiques (per nom d'arxiu original).
    en_entrada_g = set(listar_originals_rebudes(carpeta_rebudes))
    en_procesadas = set(listar_originals(carpeta_procesadas))
    en_entrada_i = set(listar_originals(carpeta_ingressos))
    en_cementiri = set(listar_originals(carpeta_retirats))
    en_lots = set()
    for rel in ("rebudes/lotes_escaneados", "rebudes/lotes_procesados",
                "apartados/lotes_vendes_escaneados", "apartados/lotes_vendes_procesados"):
        en_lots |= set(listar_originals(os.path.join(carpeta_cliente, *rel.split("/"))))

    # ------------------------------------------------------------------
    # CAPA FÍSICA
    # ------------------------------------------------------------------
    ubicacions = {
        "entrada de compres": en_entrada_g, "rebudes/procesadas": en_procesadas,
        "entrada de vendes": en_entrada_i, "errors_retirats": en_cementiri,
    }
    tots_els_noms = set().union(*ubicacions.values())
    for nombre in sorted(tots_els_noms):
        llocs = [etiqueta for etiqueta, conjunt in ubicacions.items() if nombre in conjunt]
        if len(llocs) > 1:
            divergencies.append(divergencia(
                nombre, "FÍSICA", "una sola ubicació", "present a " + " i a ".join(llocs),
            ))

    # Retirats/reactivats segons l'ÚLTIMA fila del registre de cada arxiu.
    ultima_fila_registre = {}
    for fila in registre_retirats:
        if fila.get("arxiu"):
            ultima_fila_registre[nfc(fila["arxiu"])] = fila
    for nombre, fila in sorted(ultima_fila_registre.items()):
        reactivat = (fila.get("motiu") or "").startswith("reactivat")
        base = os.path.splitext(nombre)[0]
        te_fitxa = (base + ".json") in validadas_g or (base + ".json") in validadas_i
        if not reactivat and nombre not in en_cementiri:
            divergencies.append(divergencia(
                nombre, "FÍSICA", "retirat -> a errors_retirats/",
                "no hi és (mogut o esborrat sense fila de reactivació al registre)",
            ))
        if reactivat and nombre in en_cementiri:
            divergencies.append(divergencia(
                nombre, "FÍSICA", "reactivat -> fora d'errors_retirats/",
                "encara al cementiri (la reversió no va moure l'arxiu)",
            ))
        if reactivat and nombre not in tots_els_noms and not te_fitxa:
            divergencies.append(divergencia(
                nombre, "FÍSICA", "reactivat -> present a la seva carpeta de flux",
                "no és enlloc (la resurrecció no va aterrar)",
            ))

    # ------------------------------------------------------------------
    # CAPA FITXA
    # ------------------------------------------------------------------
    for nombre_json, datos in sorted(list(validadas_g.items()) + list(validadas_i.items())):
        es_de_gastos = nombre_json in validadas_g
        if datos.get("origen") == "manual":
            continue  # arxiu nul legítim (Piso 13X/14)
        carpeta_origen = carpeta_rebudes if es_de_gastos else carpeta_ingressos
        if encontrar_original(carpeta_origen, nombre_json) is None:
            divergencies.append(divergencia(
                nombre_json, "FITXA", "fitxa activa amb original localitzable",
                f"òrfena: cap original a {os.path.relpath(carpeta_origen, RAIZ)} ni subcarpetes",
            ))
    for nombre_json in sorted((extraidas_g - set(validadas_g)) | (extraidas_i - set(validadas_i))):
        divergencies.append(divergencia(
            nombre_json, "FITXA", "tota fitxa extreta té veredicte a validadas",
            "extreta sense validar -- falta Recalcular",
        ))

    # Moviments: l'ÚLTIM per arxiu ha d'estar consumat (res abandonat a
    # l'origen; alguna cosa real al destí).
    ultim_moviment = {}
    for fila in moviments:
        if fila.get("arxiu"):
            ultim_moviment[nfc(fila["arxiu"])] = fila
    for base, fila in sorted(ultim_moviment.items()):
        de, a = fila.get("de") or "", fila.get("a") or ""
        if ":" in a:
            carpeta_desti, flux_desti = a.split(":", 1)
        else:
            carpeta_desti, flux_desti = carpeta, a
        if ":" in de:
            carpeta_de, flux_de = de.split(":", 1)
        else:
            carpeta_de, flux_de = carpeta, de
        if carpeta_de != carpeta:
            continue  # el registre del client destí l'auditarà el seu torn
        origen_flux = carpeta_rebudes if flux_de == "rebudes" else carpeta_ingressos
        if flux_de and encontrar_original(origen_flux, base + ".json"):
            # l'original segueix (tambe) a l'origen del moviment
            if carpeta_desti != carpeta or flux_desti != flux_de:
                divergencies.append(divergencia(
                    base, "FITXA", f"mogut {de} -> {a}: res abandonat a l'origen",
                    f"l'original encara és al flux {flux_de} d'aquest client",
                ))
        cc_desti = os.path.join(RAIZ, "clientes", carpeta_desti)
        if os.path.isdir(cc_desti):
            rel_i = RUTAS_ORIGEN_INGRESSOS_PERSONALIZADAS.get(carpeta_desti, "apartados/ingressos")
            origen_desti = (
                os.path.join(cc_desti, "rebudes") if flux_desti == "rebudes"
                else os.path.join(cc_desti, *rel_i.split("/"))
            )
            validades_desti = (
                os.path.join(cc_desti, "rebudes", "validadas") if flux_desti == "rebudes"
                else os.path.join(cc_desti, "apartados", "ingressos_validadas")
            )
            extraides_desti = (
                os.path.join(cc_desti, "rebudes", "extraidas") if flux_desti == "rebudes"
                else os.path.join(cc_desti, "apartados", "ingressos_extraidas")
            )
            hi_ha_algo = (
                encontrar_original(origen_desti, base + ".json")
                or os.path.exists(os.path.join(validades_desti, base + ".json"))
                or os.path.exists(os.path.join(extraides_desti, base + ".json"))
            )
            if not hi_ha_algo:
                divergencies.append(divergencia(
                    base, "FITXA", f"mogut {de} -> {a}: alguna cosa real al destí",
                    f"ni original ni fitxa a {carpeta_desti}:{flux_desti}",
                ))

    # ------------------------------------------------------------------
    # CAPA LLIBRES
    # ------------------------------------------------------------------
    detalls_destruccions = " ; ".join(
        fila.get("detall") or "" for fila in leer_csv(os.path.join(carpeta_cliente, "registre_destruccions.csv"))
    )
    for fila in hashos:
        nombre = nfc(fila.get("nombre"))
        # cargar_indice_hashos indexa TOT fitxer de la carpeta -- els
        # residus de Finder (.DS_Store) no són documents i mai tindran
        # estat; s'ignoren aquí, mai una divergència de veritat.
        if not nombre or not nombre.lower().endswith(EXTENSIONES_ORIGINAL):
            continue
        base = os.path.splitext(nombre)[0]
        existeix = (
            nombre in tots_els_noms or nombre in en_lots
            or (base + ".json") in validadas_g or (base + ".json") in validadas_i
        )
        if existeix:
            continue
        moviment = ultim_moviment.get(base)
        mogut_fora = moviment and ":" in (moviment.get("a") or "")
        destruit = base and base in detalls_destruccions
        if not mogut_fora and not destruit:
            divergencies.append(divergencia(
                nombre, "LLIBRES", "hashos.csv resol a un arxiu real o a una mort explicada",
                "el nom no és enlloc i cap llibre (moviments/destruccions) ho explica",
            ))

    for archivo, decision in sorted(decisiones.items()):
        if decision.get("accion") == "destruir":
            continue  # la fitxa es va esborrar legítimament amb certificat (13N)
        if archivo not in validadas_g and archivo not in validadas_i:
            divergencies.append(divergencia(
                archivo, "LLIBRES", "decisions.csv apunta a fitxes que existeixen",
                f"decisió '{decision.get('accion')}' sobre una fitxa que no és a validadas",
            ))

    # ------------------------------------------------------------------
    # CAPA VISTES
    # ------------------------------------------------------------------
    ruta_xlsx = os.path.join(carpeta_cliente, "sumatorios_2026.xlsx")
    hi_ha_fitxes = bool(validadas_g or validadas_i)
    if hi_ha_fitxes and not os.path.exists(ruta_xlsx):
        divergencies.append(divergencia(
            "(client)", "VISTES", "Excel generat per a un client amb fitxes", "falta Recalcular (no hi ha Excel)",
        ))
    elif hi_ha_fitxes:
        if data_mutacio_mes_nova(carpeta_cliente) > os.path.getmtime(ruta_xlsx):
            # L'Excel és més vell que l'última mutació: el detall seria
            # tot soroll amb una sola causa. El semàfor ja ho persegueix.
            divergencies.append(divergencia(
                "(client)", "VISTES", "Excel al dia amb els llibres",
                "hi ha mutacions posteriors a l'últim Recalcular -- falta Recalcular",
            ))
        else:
            divergencies += comparar_vistes(carpeta, ruta_xlsx, validadas_g, validadas_i, decisiones)

    return divergencies


def comparar_vistes(carpeta, ruta_xlsx, validadas_g, validadas_i, decisiones):
    """Conciliació de PERTINENÇA (els imports ja els vigila
    verificar_conservacio): cada fitxa viva al bloc del seu estat, cada
    morta fora. Les files d'ENTRADA MANUAL no tenen enllaç -- es
    comparen per recompte, mateix criteri que sumar/informe."""
    divergencies = []
    esperat = {}  # trimestre -> {"sumen": set(bases), "pendents": ..., "descartats": ...}
    esperat_manual = {}  # trimestre -> {"sumen": n, "pendents": n, "descartats": n}
    for nombre_json, datos in list(validadas_g.items()) + list(validadas_i.items()):
        t = trimestre_de(datos.get("fecha_factura"))
        if t is None:
            continue  # validar.py ja n'ha fet REVISAR amb motiu
        decision = decisiones.get(nombre_json)
        if estat_efectiu(decision) == "descartar":
            bloc = "descartats"
        elif datos.get("estado") == "OK" or estat_efectiu(decision) == "aprovar":
            bloc = "sumen"
        else:
            bloc = "pendents"
        if datos.get("origen") == "manual":
            comptes = esperat_manual.setdefault(t, {"sumen": 0, "pendents": 0, "descartats": 0})
            comptes[bloc] += 1
        else:
            blocs = esperat.setdefault(t, {"sumen": set(), "pendents": set(), "descartats": set()})
            blocs[bloc].add(os.path.splitext(nombre_json)[0])

    wb = load_workbook(ruta_xlsx)
    trobat = {}
    trobat_manual = {}
    for ws in wb.worksheets:
        if ws.title not in set(list(esperat) + list(esperat_manual)) and ws.title != "AVISOS":
            # full de trimestre sense fitxes esperades: si té enllaços
            # de detall, sortiran com a sobrants més avall igualment.
            pass
        if ws.title == "AVISOS":
            continue
        blocs = trobat.setdefault(ws.title, {"sumen": set(), "pendents": set(), "descartats": set()})
        comptes = trobat_manual.setdefault(ws.title, {"sumen": 0, "pendents": 0, "descartats": 0})
        seccio = None
        for row in ws.iter_rows():
            etiqueta = row[0].value
            if etiqueta in SECCIONS_DETALL:
                seccio = SECCIONS_DETALL[etiqueta]
                continue
            if etiqueta in SECCIONS_QUE_TANQUEN:
                seccio = None
                continue
            if seccio is None:
                continue
            if any(isinstance(c.value, str) and "ENTRADA MANUAL" in c.value for c in row):
                comptes[seccio] += 1
            for cell in row:
                if cell.hyperlink is not None:
                    # el target és una URI file:// amb %20 i companyia --
                    # es descodifica per comparar amb noms reals.
                    base = nfc(os.path.splitext(os.path.basename(unquote(cell.hyperlink.target)))[0])
                    blocs[seccio].add(base)
                    break

    for t in sorted(set(list(esperat) + list(trobat))):
        blocs_esp = esperat.get(t, {"sumen": set(), "pendents": set(), "descartats": set()})
        blocs_tro = trobat.get(t, {"sumen": set(), "pendents": set(), "descartats": set()})
        if t not in trobat and any(blocs_esp.values()):
            divergencies.append(divergencia(
                f"({t})", "VISTES", "full de trimestre amb el detall de les seves fitxes", "el full no existeix",
            ))
            continue
        for bloc in ("sumen", "pendents", "descartats"):
            for base in sorted(blocs_esp[bloc] - blocs_tro[bloc]):
                divergencies.append(divergencia(
                    base, "VISTES", f"{t}: al bloc '{bloc}' de l'Excel", "no hi surt",
                ))
            for base in sorted(blocs_tro[bloc] - blocs_esp[bloc]):
                altres = [b for b in ("sumen", "pendents", "descartats") if base in blocs_esp[b]]
                divergencies.append(divergencia(
                    base, "VISTES", f"{t}: fora del bloc '{bloc}'",
                    f"hi surt (li tocaria '{altres[0]}')" if altres else "hi surt i cap fitxa viva ho explica",
                ))
        comptes_esp = esperat_manual.get(t, {"sumen": 0, "pendents": 0, "descartats": 0})
        comptes_tro = trobat_manual.get(t, {"sumen": 0, "pendents": 0, "descartats": 0})
        for bloc in ("sumen", "pendents", "descartats"):
            if comptes_esp[bloc] != comptes_tro[bloc]:
                divergencies.append(divergencia(
                    f"({t} entrades manuals)", "VISTES",
                    f"{comptes_esp[bloc]} files d'ENTRADA MANUAL al bloc '{bloc}'",
                    f"{comptes_tro[bloc]}",
                ))
    return divergencies


def main():
    args = [a for a in sys.argv[1:] if a != "--informatiu"]
    informatiu = "--informatiu" in sys.argv[1:]
    carpetas = args or [f["carpeta"] for f in leer_clientes()]

    total = 0
    for carpeta in carpetas:
        if not os.path.isdir(os.path.join(RAIZ, "clientes", carpeta)):
            print(f"{carpeta}: no existeix, es salta.")
            continue
        divergencies = auditar_client(carpeta)
        total += len(divergencies)
        if not divergencies:
            print(f"{carpeta} / coherència: ✓ 0 divergències")
        for d in divergencies:
            prefix = "AVISO: coherència" if informatiu else "  ✗"
            print(f"{prefix} [{carpeta}] {d['document']} ({d['capa']}): esperat {d['esperat']} -- trobat {d['trobat']}")

    print()
    if total == 0:
        print("Coherència: ✓ 0 divergències")
    else:
        linia = f"Coherència: ✗ {total} divergències"
        print(f"AVISO: {linia}" if informatiu else linia)
    sys.exit(0 if (total == 0 or informatiu) else 1)


if __name__ == "__main__":
    main()
